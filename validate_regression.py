"""
LVM-AI Regressie Analyse – Continue LVM schatting bij atleten
- Spearman correlatie, gedeelde assen per eenheid, LVH-kleurcodering
- Bland-Altman in absolute (g) en relatieve (%) eenheden
- OLS lineaire kalibratie als primair (met geslacht conform Khurshid et al.),
  Passing-Bablok als sensitivity
- Sign-conventie: PREDICTED - CMR (intuitief: positief = AI overschat, negatief = AI onderschat)
  Dit is consistent met Khurshid et al.: negatieve waarden = CMR groter dan AI = underestimation
- 95% CI rond rho en MAE in scatter-titels (bootstrap, 1000 iter)
- Legenda overal rechts-onder, alle figuren openen in een keer

Refactored: de volledige analyse-pipeline wordt drie keer uitgevoerd:
  1. Totale cohort  -> resultaten/Totaal/
  2. Mannen (sex=1) -> resultaten/Mannen/
  3. Vrouwen (sex=0)-> resultaten/Vrouwen/

Gebruik: python validate_regression.py
"""

import subprocess, sys

def install(packages):
    for pkg in packages:
        try:
            __import__(pkg.split("==")[0].replace("-","_").replace("beautifulsoup4","bs4"))
        except ImportError:
            print(f"Installeren: {pkg}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

install([
    "tensorflow==2.19.0",
    "beautifulsoup4", "lxml",
    "numpy", "pandas", "scikit-learn",
    "matplotlib", "seaborn", "scipy",
])

import os
import base64
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from scipy import stats
from sklearn.linear_model import LinearRegression

warnings.filterwarnings("ignore")

# --- CONFIGURATIE ---------------------------------------------------------
MODEL_FILE = "ecg_rest_raw_age_sex_bmi_lvm_asymmetric_loss.h5"
ECG_DIR    = "ecg_bestanden"
LABELS_CSV = "regressionlabels.csv"
BASE_OUTPUT_DIR = "resultaten"
os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)

ECG_NORM = 2000.0
AGE_MEAN, AGE_STD = 63.35798891483556, 7.554638350423902
BMI_MEAN, BMI_STD = 27.3397, 4.7721
LVM_MEAN, LVM_STD = 89.70372484725051, 24.803669503436304
LEAD_ORDER  = ["I","II","III","V1","V2","V3","V4","V5","V6","aVF","aVL","aVR"]
ECG_SAMPLES = 5000
LVH_COLOR_MAP = [(1, "crimson", "LVH+"), (0, "steelblue", "LVH-")]
MARGIN = 0.05
LEGEND_LOC = "lower right"

def boxplot_compat(ax, data, labels, **kwargs):
    try:    return ax.boxplot(data, tick_labels=labels, **kwargs)
    except TypeError: return ax.boxplot(data, labels=labels, **kwargs)

def data_lims(*arrays, margin=MARGIN):
    vals = np.concatenate([np.asarray(a, dtype=float).ravel() for a in arrays])
    vals = vals[np.isfinite(vals)]
    if len(vals) == 0: return 0, 1
    lo, hi = vals.min(), vals.max()
    pad = (hi - lo) * margin if hi > lo else 1.0
    return lo - pad, hi + pad

def pct_diffs(pred, truth):
    """Relatieve verschillen in %: 100*(pred - truth)/mean(pred, truth).
       Positief = AI overschat, negatief = AI onderschat (Khurshid-conventie)."""
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth)
    pred, truth = pred[mask], truth[mask]
    mean_v = (pred + truth) / 2
    safe = np.abs(mean_v) > 1e-9
    return 100.0 * (pred[safe] - truth[safe]) / mean_v[safe]

def passing_bablok(x, y):
    """Passing-Bablok regression: non-parametrische method-comparison.
       Retourneert (slope, intercept). O(n^2)."""
    x = np.asarray(x, dtype=float); y = np.asarray(y, dtype=float)
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    n = len(x)
    if n < 3:
        return np.nan, np.nan
    slopes = []
    for i in range(n - 1):
        for j in range(i + 1, n):
            if x[i] != x[j]:
                s = (y[j] - y[i]) / (x[j] - x[i])
                if s != -1:
                    slopes.append(s)
    slopes = np.sort(np.array(slopes, dtype=float))
    N = len(slopes)
    if N == 0:
        return np.nan, np.nan
    K = int(np.sum(slopes < -1))
    if N % 2 == 1:
        idx = (N - 1) // 2 + K
        idx = max(0, min(idx, N - 1))
        slope = float(slopes[idx])
    else:
        i1 = max(0, min(N // 2 - 1 + K, N - 1))
        i2 = max(0, min(N // 2 + K,     N - 1))
        slope = 0.5 * float(slopes[i1] + slopes[i2])
    intercept = float(np.median(y - slope * x))
    return slope, intercept

# --- MODEL LADEN ----------------------------------------------------------
print("Model laden...")
import tensorflow as tf
model = tf.keras.models.load_model(MODEL_FILE, compile=False)
print(f"  Model geladen: {MODEL_FILE}")
print(f"  Outputs: {[o.shape for o in model.outputs]}")

# --- LABELS LADEN ---------------------------------------------------------
labels_df = pd.read_csv(LABELS_CSV, sep=";", dtype={"sample_id": str})
labels_df["sample_id"] = labels_df["sample_id"].str.strip()
labels_df.columns = [c.strip() for c in labels_df.columns]

lvh_col = next((c for c in ["lvh_label","LVH_label","LVH","lvh","lvh_status","EILVH","eilvh"]
                if c in labels_df.columns), None)

for col in ["age","sex","bmi","lvm_grams","LVM_dubBSA"]:
    if col in labels_df.columns:
        labels_df[col] = labels_df[col].astype(str).str.replace(",", ".").astype(float)

def parse_lvh(x):
    if pd.isna(x): return np.nan
    s = str(x).strip().lower()
    if s in {"1","1.0","ja","yes","y","true","t","lvh","lvh+","positive","pos"}: return 1
    if s in {"0","0.0","nee","no","n","false","f","geen","geen lvh","geen_lvh",
             "no lvh","no_lvh","lvh-","negative","neg","control"}: return 0
    try:
        v = float(s.replace(",", "."))
        if v in (0, 1): return int(v)
    except ValueError: pass
    return np.nan

has_indexed = "LVM_dubBSA" in labels_df.columns
has_lvh     = lvh_col is not None
if has_lvh:
    labels_df["lvh_label_int"] = labels_df[lvh_col].apply(parse_lvh)

print(f"\nLabels geladen: {len(labels_df)} atleten")
print(f"  CMR LVM      : {labels_df['lvm_grams'].mean():.1f} +/- {labels_df['lvm_grams'].std():.1f} g")
if has_indexed:
    print(f"  CMR LVM/BSA  : {labels_df['LVM_dubBSA'].mean():.1f} +/- {labels_df['LVM_dubBSA'].std():.1f} g/m^2")
if has_lvh:
    print(f"  LVH-kolom    : '{lvh_col}', verdeling {labels_df['lvh_label_int'].value_counts(dropna=False).to_dict()}")

# --- XML LOADER -----------------------------------------------------------
def decode_b64(raw, scale=1.0):
    return np.frombuffer(base64.b64decode(raw), dtype="<i2").astype(np.float32) * scale

def load_xml(path):
    import bs4
    with open(path, "r", errors="replace") as f:
        soup = bs4.BeautifulSoup(f, "lxml")
    v = {}
    for wf in soup.find_all("waveform"):
        wt = wf.find("waveformtype")
        if wt is None or wt.text.strip() != "Rhythm": continue
        for ld in wf.find_all("leaddata"):
            lid = ld.find("leadid").text.strip()
            sc_tag = ld.find("leadamplitudeunitsperbit")
            sc = float(sc_tag.text.replace(",", ".")) if sc_tag else 1.0
            v[lid] = decode_b64(ld.find("waveformdata").text.strip(), sc)
        break
    if "III" not in v: v["III"] = v["II"] - v["I"]
    if "aVR" not in v: v["aVR"] = -(v["I"] + v["II"]) / 2
    if "aVL" not in v: v["aVL"] =  v["I"] - v["II"] / 2
    if "aVF" not in v: v["aVF"] =  v["II"] - v["I"] / 2
    def resample(x):
        n = len(x)
        return x if n == ECG_SAMPLES else np.interp(np.linspace(0, n, ECG_SAMPLES), np.arange(n), x)
    return np.column_stack([resample(v[l]) for l in LEAD_ORDER]).astype(np.float32)

# --- INFERENTIE -----------------------------------------------------------
print("\nInferentie starten...")
ecg_index = {Path(p).stem: str(p) for p in Path(ECG_DIR).glob("*.xml")}
print(f"  ECG-bestanden gevonden: {len(ecg_index)}")

n_inputs = len(model.inputs)
records, fouten = [], 0
for i, row in labels_df.iterrows():
    sid = str(row["sample_id"])
    if sid not in ecg_index: fouten += 1; continue
    try:
        ecg  = load_xml(ecg_index[sid])
        norm = (ecg / ECG_NORM)[np.newaxis, ...]
        age_n = np.array([[(float(row["age"]) - AGE_MEAN) / AGE_STD]], dtype=np.float32)
        bmi_n = np.array([[(float(row["bmi"]) - BMI_MEAN) / BMI_STD]], dtype=np.float32)
        sex_val = int(row["sex"])
        sex = np.array([[1 - sex_val, sex_val]], dtype=np.float32)

        if n_inputs == 4:   out = model.predict([norm, age_n, sex, bmi_n], verbose=0)
        elif n_inputs == 2: out = model.predict([norm, np.array([[age_n[0,0], sex_val, bmi_n[0,0]]], dtype=np.float32)], verbose=0)
        else:               out = model.predict(norm, verbose=0)

        reg_raw = None
        if isinstance(out, (list, tuple)):
            for o in out:
                if o.shape[-1] == 1:
                    reg_raw = float(o[0, 0]); break
        else:
            reg_raw = float(out[0, 0])
        if reg_raw is None: fouten += 1; continue

        rec = {
            "sample_id":         sid,
            "age":               float(row["age"]),
            "sex":               sex_val,
            "bmi":               float(row["bmi"]),
            "lvm_cmr":           float(row["lvm_grams"]),
            "lvm_predicted":     reg_raw * LVM_STD + LVM_MEAN,
            "lvm_predicted_raw": reg_raw,
        }
        if has_indexed:
            rec["lvm_cmr_indexed"]       = float(row["LVM_dubBSA"])
            bsa = rec["lvm_cmr"] / rec["lvm_cmr_indexed"] if rec["lvm_cmr_indexed"] else np.nan
            rec["bsa"]                   = bsa
            rec["lvm_predicted_indexed"] = rec["lvm_predicted"] / bsa if bsa else np.nan
        if has_lvh:
            rec["lvh_label_int"] = row.get("lvh_label_int", np.nan)
        records.append(rec)
        if (i + 1) % 50 == 0: print(f"  {i+1}/{len(labels_df)} verwerkt...")
    except Exception as e:
        print(f"  FOUT bij {sid}: {e}"); fouten += 1

full_results = pd.DataFrame(records)
print(f"\nKlaar: {len(full_results)} voorspellingen, {fouten} fouten")
if len(full_results) == 0: sys.exit("Geen voorspellingen gelukt.")
full_results = full_results[np.isfinite(full_results["lvm_predicted"]) & np.isfinite(full_results["lvm_cmr"])].reset_index(drop=True)
if len(full_results) < 3: sys.exit("Te weinig geldige voorspellingen.")

# --- METRIC HELPERS -------------------------------------------------------
def metrics(pred, truth, label, seed=42):
    """Compute metrics with Predicted - CMR sign convention.
       Positive bias = AI overestimates; negative bias = AI underestimates."""
    pred  = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask  = np.isfinite(pred) & np.isfinite(truth); pred, truth = pred[mask], truth[mask]
    if len(pred) < 3:
        return dict(label=label, n=len(pred), rho=np.nan, p=np.nan, rho_lo=np.nan, rho_hi=np.nan,
                    mae=np.nan, m_lo=np.nan, m_hi=np.nan, mean_d=np.nan, sd_d=np.nan,
                    loa_lo=np.nan, loa_hi=np.nan,
                    smape=np.nan, bias_pct=np.nan, loa_pct_lo=np.nan, loa_pct_hi=np.nan)
    rho, p_val = stats.spearmanr(pred, truth)
    diff = pred - truth
    mae = np.mean(np.abs(diff)); mean_d, sd_d = diff.mean(), diff.std()
    mean_v = (pred + truth) / 2
    safe   = np.abs(mean_v) > 1e-9
    diff_pct = 100.0 * (pred[safe] - truth[safe]) / mean_v[safe]
    smape    = float(np.mean(np.abs(diff_pct)))
    bias_pct = float(diff_pct.mean()); sd_pct = float(diff_pct.std())
    rng = np.random.default_rng(seed); boot_rho, boot_m = [], []
    for _ in range(1000):
        idx = rng.integers(0, len(truth), len(truth))
        boot_rho.append(stats.spearmanr(pred[idx], truth[idx])[0])
        boot_m.append(np.mean(np.abs(pred[idx] - truth[idx])))
    rho_lo, rho_hi = np.percentile(boot_rho, [2.5, 97.5])
    m_lo, m_hi     = np.percentile(boot_m,   [2.5, 97.5])
    return dict(label=label, n=len(pred), rho=rho, p=p_val, rho_lo=rho_lo, rho_hi=rho_hi,
                mae=mae, m_lo=m_lo, m_hi=m_hi, mean_d=mean_d, sd_d=sd_d,
                loa_lo=mean_d - 1.96*sd_d, loa_hi=mean_d + 1.96*sd_d,
                smape=smape, bias_pct=bias_pct,
                loa_pct_lo=bias_pct - 1.96*sd_pct,
                loa_pct_hi=bias_pct + 1.96*sd_pct)

def print_metrics(m):
    print(f"\n  {m['label']} (n={m['n']})")
    if np.isnan(m['rho']): print(f"    (te weinig data)"); return
    print(f"    Spearman rho            : {m['rho']:.3f}  95%CI [{m['rho_lo']:.3f}, {m['rho_hi']:.3f}]   p={m['p']:.2e}")
    print(f"    MAE (absoluut)          : {m['mae']:.1f}  95%CI [{m['m_lo']:.1f}, {m['m_hi']:.1f}]")
    print(f"    sMAPE (relatief)        : {m['smape']:.1f}%")
    print(f"    Bias absoluut (Pred-CMR): {m['mean_d']:+.1f}      LoA [{m['loa_lo']:+.1f}, {m['loa_hi']:+.1f}]")
    print(f"    Bias relatief (Pred-CMR): {m['bias_pct']:+.1f}%   LoA [{m['loa_pct_lo']:+.1f}%, {m['loa_pct_hi']:+.1f}%]")
    print(f"    (positief = AI overschat, negatief = AI onderschat)")

def _split_by_color(x, y, color_by, color_map):
    cb = np.asarray(color_by); groups = []; plotted = np.zeros(len(cb), dtype=bool)
    for val, col, lbl in color_map:
        sel = cb == val; plotted |= sel
        if sel.any(): groups.append((x[sel], y[sel], col, lbl))
    unk = ~plotted
    if unk.any(): groups.append((x[unk], y[unk], "lightgray", f"Onbekend (n={unk.sum()})"))
    return groups

def scatter_panel(ax, pred, truth, title, xlim, ylim,
                  color_by=None, color_map=None, seed=42):
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth); pred, truth = pred[mask], truth[mask]
    if color_by is not None and color_map is not None:
        cb = np.asarray(color_by)[mask]
        for x_, y_, col, lbl in _split_by_color(pred, truth, cb, color_map):
            ax.scatter(x_, y_, alpha=0.55, s=20, color=col, label=lbl)
    else:
        ax.scatter(pred, truth, alpha=0.4, s=20, color="royalblue")
    lo_lim = min(xlim[0], ylim[0]); hi_lim = max(xlim[1], ylim[1])
    ax.plot([lo_lim, hi_lim], [lo_lim, hi_lim], "--", color="gray", label="Identity")
    if len(pred) >= 3:
        rho_val, _ = stats.spearmanr(pred, truth)
        mae_val = float(np.mean(np.abs(pred - truth)))
        rng = np.random.default_rng(seed); boot_rho, boot_mae = [], []
        for _ in range(1000):
            idx = rng.integers(0, len(truth), len(truth))
            boot_rho.append(stats.spearmanr(pred[idx], truth[idx])[0])
            boot_mae.append(np.mean(np.abs(pred[idx] - truth[idx])))
        rho_lo, rho_hi = np.percentile(boot_rho, [2.5, 97.5])
        mae_lo, mae_hi = np.percentile(boot_mae, [2.5, 97.5])
        ax.set_title(
            f"{title}\n"
            f"rho={rho_val:.2f} (95% CI {rho_lo:.2f}-{rho_hi:.2f}), "
            f"MAE={mae_val:.1f} (95% CI {mae_lo:.1f}-{mae_hi:.1f})",
            fontsize=10
        )
    else:
        ax.set_title(title)
    ax.set_xlabel("Predicted"); ax.set_ylabel("CMR-derived")
    ax.set_xlim(xlim); ax.set_ylim(ylim)
    ax.legend(fontsize=8, loc=LEGEND_LOC)

def ba_panel(ax, pred, truth, title, xlim_ba, ylim_ba, color_by=None, color_map=None):
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth); pred, truth = pred[mask], truth[mask]
    if len(pred) == 0:
        ax.set_title(title); ax.text(0.5, 0.5, "Geen data", ha="center", transform=ax.transAxes); return
    mean_v = (pred + truth) / 2
    diff_v = pred - truth
    md, sd = diff_v.mean(), diff_v.std()
    if color_by is not None and color_map is not None:
        cb = np.asarray(color_by)[mask]
        for x_, y_, col, lbl in _split_by_color(mean_v, diff_v, cb, color_map):
            ax.scatter(x_, y_, alpha=0.55, s=20, color=col, label=lbl)
    else:
        ax.scatter(mean_v, diff_v, alpha=0.4, s=20, color="royalblue")
    ax.axhline(y=md,           color="red", linestyle="-",  label=f"Bias {md:+.1f}")
    ax.axhline(y=md + 1.96*sd, color="red", linestyle="--", label=f"+1.96 SD {md+1.96*sd:+.1f}")
    ax.axhline(y=md - 1.96*sd, color="red", linestyle="--", label=f"-1.96 SD {md-1.96*sd:+.1f}")
    ax.axhline(y=0, color="gray", linestyle=":")
    ax.set_xlabel("Mean of CMR and predicted")
    ax.set_ylabel("Predicted - CMR")
    ax.set_title(f"BA absoluut: {title}")
    ax.set_xlim(xlim_ba); ax.set_ylim(ylim_ba)
    ax.legend(fontsize=8, loc=LEGEND_LOC)

def ba_panel_relative(ax, pred, truth, title, xlim_ba, ylim_ba_pct,
                      color_by=None, color_map=None):
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth); pred, truth = pred[mask], truth[mask]
    if color_by is not None: color_by = np.asarray(color_by)[mask]
    if len(pred) == 0:
        ax.set_title(title); ax.text(0.5, 0.5, "Geen data", ha="center", transform=ax.transAxes); return
    mean_v = (pred + truth) / 2
    safe   = np.abs(mean_v) > 1e-9
    mean_v = mean_v[safe]; pred = pred[safe]; truth = truth[safe]
    if color_by is not None: color_by = color_by[safe]
    diff_pct = 100.0 * (pred - truth) / mean_v
    md, sd = diff_pct.mean(), diff_pct.std()
    if color_by is not None and color_map is not None:
        for x_, y_, col, lbl in _split_by_color(mean_v, diff_pct, color_by, color_map):
            ax.scatter(x_, y_, alpha=0.55, s=20, color=col, label=lbl)
    else:
        ax.scatter(mean_v, diff_pct, alpha=0.4, s=20, color="royalblue")
    ax.axhline(y=md,           color="red", linestyle="-",  label=f"Bias {md:+.1f}%")
    ax.axhline(y=md + 1.96*sd, color="red", linestyle="--", label=f"+1.96 SD {md+1.96*sd:+.1f}%")
    ax.axhline(y=md - 1.96*sd, color="red", linestyle="--", label=f"-1.96 SD {md-1.96*sd:+.1f}%")
    ax.axhline(y=0, color="gray", linestyle=":")
    ax.set_xlabel("Mean of CMR and predicted")
    ax.set_ylabel("(Predicted - CMR) / Mean * 100 (%)")
    ax.set_title(f"BA relatief: {title}")
    ax.set_xlim(xlim_ba); ax.set_ylim(ylim_ba_pct)
    ax.legend(fontsize=8, loc=LEGEND_LOC)

def stats_box(ax, pred, truth, seed=42):
    """Bottom-right stats box: Spearman r [CI]; p, and MAE [CI]."""
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth)
    pred, truth = pred[mask], truth[mask]
    if len(pred) < 3:
        return
    rho_val, p_val = stats.spearmanr(pred, truth)
    mae_val = float(np.mean(np.abs(pred - truth)))
    rng = np.random.default_rng(seed)
    boot_rho, boot_mae = [], []
    for _ in range(1000):
        idx = rng.integers(0, len(truth), len(truth))
        boot_rho.append(stats.spearmanr(pred[idx], truth[idx])[0])
        boot_mae.append(np.mean(np.abs(pred[idx] - truth[idx])))
    rho_lo, rho_hi = np.percentile(boot_rho, [2.5, 97.5])
    mae_lo, mae_hi = np.percentile(boot_mae, [2.5, 97.5])
    if p_val < 0.001:
        p_str = "p < 0.001"
    else:
        p_str = f"p = {p_val:.3f}"
    txt = (f"Spearman ρ = {rho_val:.2f} (95% CI {rho_lo:.2f}-{rho_hi:.2f}); {p_str}\n"
           f"MAE = {mae_val:.1f} (95% CI {mae_lo:.1f}-{mae_hi:.1f})")
    ax.text(0.97, 0.03, txt, transform=ax.transAxes,
            ha="right", va="bottom", fontsize=9,
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white",
                      edgecolor="gray", alpha=0.9))

def scatter_panel_en(ax, pred, truth, xlabel, ylabel, panel_letter,
                     panel_title, xlim, ylim, color_by=None, color_map=None):
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth)
    pred_m, truth_m = pred[mask], truth[mask]
    if color_by is not None and color_map is not None:
        cb = np.asarray(color_by)[mask]
        for x_, y_, col, lbl in _split_by_color(pred_m, truth_m, cb, color_map):
            ax.scatter(x_, y_, alpha=0.55, s=22, color=col, label=lbl)
    else:
        ax.scatter(pred_m, truth_m, alpha=0.5, s=22, color="royalblue")
    lo_lim = min(xlim[0], ylim[0]); hi_lim = max(xlim[1], ylim[1])
    ax.plot([lo_lim, hi_lim], [lo_lim, hi_lim], "--", color="gray")
    ax.set_xlabel(xlabel); ax.set_ylabel(ylabel)
    ax.set_xlim(xlim); ax.set_ylim(ylim)
    ax.set_title(panel_title, fontsize=11)
    ax.text(0.02, 0.98, panel_letter, transform=ax.transAxes,
            fontsize=14, fontweight="bold", va="top", ha="left")
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(handles, labels, fontsize=8, loc="upper left",
                  bbox_to_anchor=(0.02, 0.92))
    stats_box(ax, pred, truth)

def ba_panel_centered(ax, pred, truth, panel_letter, title_text,
                      xlim, ylim_sym, color_by=None, color_map=None,
                      relative=True):
    """BA panel with zero-centered y-axis. Sign convention: Predicted - CMR."""
    pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth)
    pred, truth = pred[mask], truth[mask]
    if color_by is not None:
        color_by = np.asarray(color_by)[mask]
    if len(pred) == 0:
        ax.set_title(title_text)
        ax.text(0.5, 0.5, "No data", ha="center", transform=ax.transAxes)
        return
    mean_v = (pred + truth) / 2
    if relative:
        safe = np.abs(mean_v) > 1e-9
        mean_v = mean_v[safe]; pred = pred[safe]; truth = truth[safe]
        if color_by is not None: color_by = color_by[safe]
        diff = 100.0 * (pred - truth) / mean_v
        ylabel = "(Predicted - CMR) / Mean × 100 (%)"
        unit = "%"
    else:
        diff = pred - truth
        ylabel = "Predicted - CMR"
        unit = ""
    md, sd = diff.mean(), diff.std()
    if color_by is not None and color_map is not None:
        for x_, y_, col, lbl in _split_by_color(mean_v, diff, color_by, color_map):
            ax.scatter(x_, y_, alpha=0.6, s=22, color=col, label=lbl)
    else:
        ax.scatter(mean_v, diff, alpha=0.5, s=22, color="royalblue")
    ax.axhline(md, color="red", linestyle="-",
               label=f"Bias {md:+.1f}{unit}")
    ax.axhline(md + 1.96*sd, color="red", linestyle="--",
               label=f"+1.96 SD {md+1.96*sd:+.1f}{unit}")
    ax.axhline(md - 1.96*sd, color="red", linestyle="--",
               label=f"-1.96 SD {md-1.96*sd:+.1f}{unit}")
    ax.axhline(0, color="gray", linestyle=":")
    ax.set_xlabel("Mean of CMR and predicted")
    ax.set_ylabel(ylabel)
    ax.set_title(title_text)
    ax.set_xlim(xlim); ax.set_ylim(ylim_sym)
    ax.text(0.02, 0.98, panel_letter, transform=ax.transAxes,
            fontsize=14, fontweight="bold", va="top", ha="left")
    ax.legend(fontsize=8, loc=LEGEND_LOC)

def ba_relative_strict(ax, pred, truth, panel_letter, subplot_title,
                       xlabel, xlim, ylim_sym,
                       color_by=None, color_map=None):
    """Relative BA panel. Sign convention: Predicted - CMR."""
    pred = np.asarray(pred, dtype=float)
    truth = np.asarray(truth, dtype=float)
    mask = np.isfinite(pred) & np.isfinite(truth)
    pred, truth = pred[mask], truth[mask]
    if color_by is not None:
        color_by = np.asarray(color_by)[mask]
    if len(pred) == 0:
        ax.set_title(subplot_title)
        ax.text(0.5, 0.5, "No data", ha="center", transform=ax.transAxes)
        return None

    mean_v = (pred + truth) / 2.0
    safe = np.abs(mean_v) > 1e-9
    mean_v = mean_v[safe]
    pred_s, truth_s = pred[safe], truth[safe]
    if color_by is not None:
        color_by = color_by[safe]
    diff_pct = 100.0 * (pred_s - truth_s) / mean_v

    bias = float(diff_pct.mean())
    sd = float(diff_pct.std())
    upper = bias + 1.96 * sd
    lower = bias - 1.96 * sd

    if color_by is not None and color_map is not None:
        for x_, y_, col, lbl in _split_by_color(mean_v, diff_pct, color_by, color_map):
            ax.scatter(x_, y_, alpha=0.6, s=22, color=col, label=lbl)
    else:
        ax.scatter(mean_v, diff_pct, alpha=0.5, s=22, color="royalblue")

    ax.axhline(0,     color="grey",     linestyle="--")
    ax.axhline(bias,  color="darkblue", linestyle="-")
    ax.axhline(upper, color="red",      linestyle="--")
    ax.axhline(lower, color="red",      linestyle="--")

    ax.set_xlabel(xlabel)
    ax.set_ylabel("Relative Difference (%)")
    ax.set_title(subplot_title, fontsize=11)
    ax.set_xlim(xlim)
    ax.set_ylim(ylim_sym)

    ax.text(0.02, 0.985, f"({panel_letter})", transform=ax.transAxes,
            fontsize=14, fontweight="bold", va="top", ha="left")

    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(handles, labels, loc="upper left",
                  bbox_to_anchor=(0.02, 0.92), fontsize=8, framealpha=0.9)

    stats_txt = (f"Bias: {bias:+.1f}%\n"
                 f"Upper LoA: {upper:+.1f}%\n"
                 f"Lower LoA: {lower:+.1f}%")
    ax.text(0.97, 0.03, stats_txt, transform=ax.transAxes,
            ha="right", va="bottom", fontsize=9,
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white",
                      edgecolor="gray", alpha=0.9))

    return dict(bias=bias, upper=upper, lower=lower, n=len(diff_pct))


def fmt_p_boxplot(p):
    """Format p-value: p < 0.001 or p = 0.XXX"""
    if not np.isfinite(p):
        return ""
    if p < 0.001:
        return "p < 0.001"
    return f"p = {p:.3f}"


# --- ROC helper ----------------------------------------------------------
from sklearn.metrics import roc_curve, auc

def bootstrap_auc(y_true, y_score, n_boot=1000, seed=42):
    """Bootstrap 95% CI for AUC."""
    y_true = np.asarray(y_true)
    y_score = np.asarray(y_score)
    mask = np.isfinite(y_score) & np.isfinite(y_true)
    y_true, y_score = y_true[mask], y_score[mask]
    if len(np.unique(y_true)) < 2 or len(y_true) < 10:
        return np.nan, np.nan, np.nan
    fpr, tpr, _ = roc_curve(y_true, y_score)
    auc_val = auc(fpr, tpr)
    rng = np.random.default_rng(seed)
    boot_aucs = []
    for _ in range(n_boot):
        idx = rng.integers(0, len(y_true), len(y_true))
        if len(np.unique(y_true[idx])) < 2:
            continue
        f, t, _ = roc_curve(y_true[idx], y_score[idx])
        boot_aucs.append(auc(f, t))
    if len(boot_aucs) == 0:
        return auc_val, np.nan, np.nan
    lo, hi = np.percentile(boot_aucs, [2.5, 97.5])
    return auc_val, lo, hi


# --- Excel styling helpers -----------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

F_BODY    = Font(name="Arial", size=10)
F_BOLD    = Font(name="Arial", size=10, bold=True)
F_TITLE   = Font(name="Arial", size=11, bold=True)
F_ITALIC  = Font(name="Arial", size=9,  italic=True)
FILL_SEC  = PatternFill("solid", start_color="D9D9D9")
FILL_HDR  = PatternFill("solid", start_color="F2F2F2")
SIDE_THIN  = Side(border_style="thin",   color="808080")
SIDE_THICK = Side(border_style="medium", color="000000")
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def f_ref(values):
    v = np.asarray(values, dtype=float); v = v[np.isfinite(v)]
    return "-" if len(v) == 0 else f"{v.mean():.1f} ± {v.std():.1f}"

def f_rho(m):
    if m is None or not np.isfinite(m["rho"]): return "-"
    return f"{m['rho']:.2f} (95% CI {m['rho_lo']:.2f}-{m['rho_hi']:.2f})"

def f_mae(m):
    if m is None or not np.isfinite(m["mae"]): return "-"
    return f"{m['mae']:.1f} (95% CI {m['m_lo']:.1f}-{m['m_hi']:.1f})"

def f_bias_abs(m):
    if m is None or not np.isfinite(m["mean_d"]): return "-"
    return f"{m['mean_d']:+.1f}"

def f_loa_abs(m):
    if m is None or not (np.isfinite(m["loa_lo"]) and np.isfinite(m["loa_hi"])):
        return "-"
    return f"{m['loa_lo']:+.1f} to {m['loa_hi']:+.1f}"

def f_smape(m):
    if m is None or not np.isfinite(m["smape"]): return "-"
    return f"{m['smape']:.1f}"

def f_bias_pct(m):
    if m is None or not np.isfinite(m.get("bias_pct", np.nan)): return "-"
    return f"{m['bias_pct']:+.1f}"

def f_loa_pct(m):
    if m is None or not (np.isfinite(m.get("loa_pct_lo", np.nan))
                         and np.isfinite(m.get("loa_pct_hi", np.nan))):
        return "-"
    return f"{m['loa_pct_lo']:+.1f} to {m['loa_pct_hi']:+.1f}"

def f_p(p):
    if not np.isfinite(p): return "n/a"
    return "p < 0.001" if p < 0.001 else f"p = {p:.3f}"

def _write_sheet(ws, title, col_headers, rows, footnotes, col_widths):
    n_cols = len(col_headers) + 1
    ws.cell(row=1, column=1, value=title).font = F_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 24

    ws.cell(row=2, column=1, value="").border = Border(
        top=SIDE_THICK, bottom=SIDE_THICK, left=SIDE_THIN, right=SIDE_THIN)
    for i, h in enumerate(col_headers):
        c = ws.cell(row=2, column=2 + i, value=h)
        c.font = F_BOLD; c.alignment = A_CENTER; c.fill = FILL_HDR
        c.border = Border(top=SIDE_THICK, bottom=SIDE_THICK,
                          left=SIDE_THIN, right=SIDE_THIN)
    ws.row_dimensions[2].height = 36

    r = 3
    for label, values, is_header in rows:
        ws.cell(row=r, column=1, value=label)
        for i, v in enumerate(values):
            ws.cell(row=r, column=2 + i, value=v)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = F_BOLD if is_header else F_BODY
            cell.alignment = A_LEFT if col == 1 else A_CENTER
            if is_header:
                cell.fill = FILL_SEC
            cell.border = Border(left=SIDE_THIN, right=SIDE_THIN,
                                 top=SIDE_THIN, bottom=SIDE_THIN)
        ws.row_dimensions[r].height = 22 if is_header else 18
        r += 1

    for col in range(1, n_cols + 1):
        c = ws.cell(row=r - 1, column=col)
        c.border = Border(left=SIDE_THIN, right=SIDE_THIN,
                          top=SIDE_THIN, bottom=SIDE_THICK)

    for foot in footnotes:
        r += 1
        ws.cell(row=r, column=1, value=foot).font = F_ITALIC
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        ws.cell(row=r, column=1).alignment = A_LEFT
        ws.row_dimensions[r].height = 30 if len(foot) > 110 else 18

    ws.column_dimensions["A"].width = col_widths[0]
    for i, w in enumerate(col_widths[1:]):
        ws.column_dimensions[chr(ord("B") + i)].width = w

    ws.freeze_panes = "B3"


# =========================================================================
# === MASTER ANALYSE-FUNCTIE ==============================================
# =========================================================================
def run_analysis(results, cohort_name, output_dir):
    """Voert de volledige analyse-pipeline uit op de gegeven (gefilterde) DataFrame.

    Parameters
    ----------
    results : pd.DataFrame
        Voorspellings-DataFrame (mogelijk gefilterd op sex).
    cohort_name : str
        Naam van de cohort (bv. 'Totaal', 'Mannen', 'Vrouwen').
    output_dir : str
        Map waar alle figuren/CSV/Excel naartoe geschreven worden.
    """
    print(f"\n\n{'#'*72}")
    print(f"#  ANALYSE VOOR COHORT: {cohort_name}  (n={len(results)})")
    print(f"#  Output directory: {output_dir}")
    print(f"{'#'*72}")

    os.makedirs(output_dir, exist_ok=True)

    # --- Graceful degradation: te weinig data --------------------------
    if len(results) < 5:
        print(f"  [WAARSCHUWING] Cohort '{cohort_name}' heeft minder dan 5 rijen "
              f"(n={len(results)}). Volledige analyse overgeslagen.")
        return

    results = results.reset_index(drop=True).copy()

    # --- 1. ABSOLUTE LVM - OLS-kalibratie -----------------------------
    lvm_cmr  = results["lvm_cmr"].values
    lvm_pred = results["lvm_predicted"].values
    sex_arr  = results["sex"].values
    lvh_arr  = results["lvh_label_int"].values if has_lvh else None
    cmap     = LVH_COLOR_MAP if has_lvh else None

    # OLS met geslacht als covariaat (singular indien cohort 1 geslacht heeft;
    # sklearn lost dit op via lstsq - voorspellingen blijven correct).
    X_abs = np.column_stack([lvm_pred, sex_arr])
    calib_abs = LinearRegression().fit(X_abs, lvm_cmr)
    coef_pred_abs = float(calib_abs.coef_[0])
    coef_sex_abs  = float(calib_abs.coef_[1])
    intercept_abs = float(calib_abs.intercept_)
    lvm_pred_cal  = calib_abs.predict(X_abs)
    results["lvm_predicted_calibrated"] = lvm_pred_cal

    # Passing-Bablok sensitivity (univariaat)
    pb_slope_abs, pb_intercept_abs = passing_bablok(lvm_pred, lvm_cmr)
    lvm_pred_pb = pb_slope_abs * lvm_pred + pb_intercept_abs
    results["lvm_predicted_pb"] = lvm_pred_pb

    sc_lim1   = data_lims(lvm_cmr, lvm_pred, lvm_pred_cal, lvm_pred_pb)
    ba_x_lim1 = data_lims((lvm_pred+lvm_cmr)/2, (lvm_pred_cal+lvm_cmr)/2, (lvm_pred_pb+lvm_cmr)/2)
    ba_y_lim1 = data_lims(lvm_pred - lvm_cmr, lvm_pred_cal - lvm_cmr, lvm_pred_pb - lvm_cmr)
    ba_y_lim1_pct = data_lims(pct_diffs(lvm_pred, lvm_cmr),
                              pct_diffs(lvm_pred_cal, lvm_cmr),
                              pct_diffs(lvm_pred_pb,  lvm_cmr))

    print(f"\n{'='*64}")
    print(f"  1. ABSOLUTE LVM (g) - primary: OLS  (bias = Predicted - CMR)")
    print(f"     [positief = AI overschat, negatief = AI onderschat]")
    print(f"{'='*64}")
    print(f"  Gedeelde scatter-as    : {sc_lim1[0]:.0f} - {sc_lim1[1]:.0f} g")
    print(f"  Gedeelde BA-x-as       : {ba_x_lim1[0]:.0f} - {ba_x_lim1[1]:.0f} g")
    print(f"  Gedeelde BA-y-as abs   : {ba_y_lim1[0]:.0f} - {ba_y_lim1[1]:.0f} g")
    print(f"  Gedeelde BA-y-as %     : {ba_y_lim1_pct[0]:.1f} - {ba_y_lim1_pct[1]:.1f} %")
    print(f"  CMR LVM                : {np.mean(lvm_cmr):.1f} +/- {np.std(lvm_cmr):.1f}  (range {lvm_cmr.min():.0f}-{lvm_cmr.max():.0f})")
    print(f"  Predicted LVM (ruw)    : {np.mean(lvm_pred):.1f} +/- {np.std(lvm_pred):.1f}  (range {lvm_pred.min():.0f}-{lvm_pred.max():.0f})")
    print(f"  OLS-formule            : CMR = {coef_pred_abs:.3f} * predicted + {coef_sex_abs:+.1f} * sex + {intercept_abs:+.1f}")
    print(f"  PB-formule  (sensit.)  : CMR = {pb_slope_abs:.3f} * predicted + {pb_intercept_abs:+.1f}  (zonder geslacht)")
    print_metrics(metrics(lvm_pred,      lvm_cmr, "Ruw"))
    print_metrics(metrics(lvm_pred_cal,  lvm_cmr, "OLS-gerecalibreerd"))
    print_metrics(metrics(lvm_pred_pb,   lvm_cmr, "PB-gerecalibreerd (sensitivity)"))

    fig1, ax1 = plt.subplots(2, 3, figsize=(18, 11))
    scatter_panel(ax1[0,0], lvm_pred, lvm_cmr, "Ruw",
                  sc_lim1, sc_lim1, color_by=lvh_arr, color_map=cmap)
    ba_panel(ax1[0,1], lvm_pred, lvm_cmr, "Ruw",
             ba_x_lim1, ba_y_lim1, color_by=lvh_arr, color_map=cmap)
    ba_panel_relative(ax1[0,2], lvm_pred, lvm_cmr, "Ruw",
                      ba_x_lim1, ba_y_lim1_pct, color_by=lvh_arr, color_map=cmap)
    scatter_panel(ax1[1,0], lvm_pred_cal, lvm_cmr, "OLS-gerecalibreerd",
                  sc_lim1, sc_lim1, color_by=lvh_arr, color_map=cmap)
    ba_panel(ax1[1,1], lvm_pred_cal, lvm_cmr, "OLS-gerecalibreerd",
             ba_x_lim1, ba_y_lim1, color_by=lvh_arr, color_map=cmap)
    ba_panel_relative(ax1[1,2], lvm_pred_cal, lvm_cmr, "OLS-gerecalibreerd",
                      ba_x_lim1, ba_y_lim1_pct, color_by=lvh_arr, color_map=cmap)
    fig1.suptitle(f"[{cohort_name}] LVM-AI vs CMR - Absolute LVM (g), OLS-kalibratie  (Predicted - CMR)", fontsize=14)
    fig1.tight_layout()
    fig1.savefig(os.path.join(output_dir, "fig1_absoluut.png"), dpi=150)
    print(f"\n  Figuur 1 opgeslagen: {output_dir}/fig1_absoluut.png")

    # --- 2. GEINDEXEERDE LVM - OLS-kalibratie -------------------------
    sc_lim2 = ba_x_lim2 = ba_y_lim2 = ba_y_lim2_pct = None
    coef_pred_ix = coef_sex_ix = intercept_ix = None
    pb_slope_ix = pb_intercept_ix = None

    if has_indexed and "lvm_predicted_indexed" in results.columns:
        ix_mask = np.isfinite(results["lvm_predicted_indexed"]) & np.isfinite(results["lvm_cmr_indexed"])
        if ix_mask.sum() >= 3:
            lvm_cmr_ix  = results.loc[ix_mask, "lvm_cmr_indexed"].values
            lvm_pred_ix = results.loc[ix_mask, "lvm_predicted_indexed"].values
            sex_arr_ix  = results.loc[ix_mask, "sex"].values
            lvh_arr_ix  = results.loc[ix_mask, "lvh_label_int"].values if has_lvh else None

            X_ix = np.column_stack([lvm_pred_ix, sex_arr_ix])
            calib_ix = LinearRegression().fit(X_ix, lvm_cmr_ix)
            coef_pred_ix = float(calib_ix.coef_[0])
            coef_sex_ix  = float(calib_ix.coef_[1])
            intercept_ix = float(calib_ix.intercept_)
            lvm_pred_ix_cal = calib_ix.predict(X_ix)
            results.loc[ix_mask, "lvm_predicted_indexed_calibrated"] = lvm_pred_ix_cal

            pb_slope_ix, pb_intercept_ix = passing_bablok(lvm_pred_ix, lvm_cmr_ix)
            lvm_pred_ix_pb = pb_slope_ix * lvm_pred_ix + pb_intercept_ix
            results.loc[ix_mask, "lvm_predicted_indexed_pb"] = lvm_pred_ix_pb

            sc_lim2   = data_lims(lvm_cmr_ix, lvm_pred_ix, lvm_pred_ix_cal, lvm_pred_ix_pb)
            ba_x_lim2 = data_lims((lvm_pred_ix+lvm_cmr_ix)/2, (lvm_pred_ix_cal+lvm_cmr_ix)/2,
                                  (lvm_pred_ix_pb+lvm_cmr_ix)/2)
            ba_y_lim2 = data_lims(lvm_pred_ix - lvm_cmr_ix, lvm_pred_ix_cal - lvm_cmr_ix,
                                  lvm_pred_ix_pb - lvm_cmr_ix)
            ba_y_lim2_pct = data_lims(pct_diffs(lvm_pred_ix,     lvm_cmr_ix),
                                      pct_diffs(lvm_pred_ix_cal, lvm_cmr_ix),
                                      pct_diffs(lvm_pred_ix_pb,  lvm_cmr_ix))

            print(f"\n{'='*64}")
            print(f"  2. GEINDEXEERDE LVM (g/m^2) - primary: OLS  (bias = Predicted - CMR)")
            print(f"{'='*64}")
            print(f"  Gedeelde scatter-as    : {sc_lim2[0]:.0f} - {sc_lim2[1]:.0f} g/m^2")
            print(f"  Gedeelde BA-x-as       : {ba_x_lim2[0]:.0f} - {ba_x_lim2[1]:.0f} g/m^2")
            print(f"  Gedeelde BA-y-as abs   : {ba_y_lim2[0]:.0f} - {ba_y_lim2[1]:.0f} g/m^2")
            print(f"  Gedeelde BA-y-as %     : {ba_y_lim2_pct[0]:.1f} - {ba_y_lim2_pct[1]:.1f} %")
            print(f"  CMR LVM/BSA            : {np.mean(lvm_cmr_ix):.1f} +/- {np.std(lvm_cmr_ix):.1f}")
            print(f"  Predicted LVM/BSA (ruw): {np.mean(lvm_pred_ix):.1f} +/- {np.std(lvm_pred_ix):.1f}")
            print(f"  OLS-formule            : CMR/BSA = {coef_pred_ix:.3f} * predicted/BSA + {coef_sex_ix:+.1f} * sex + {intercept_ix:+.1f}")
            print(f"  PB-formule  (sensit.)  : CMR/BSA = {pb_slope_ix:.3f} * predicted/BSA + {pb_intercept_ix:+.1f}  (zonder geslacht)")
            print_metrics(metrics(lvm_pred_ix,     lvm_cmr_ix, "Geindexeerd ruw"))
            print_metrics(metrics(lvm_pred_ix_cal, lvm_cmr_ix, "Geindexeerd OLS-gerecalibreerd"))
            print_metrics(metrics(lvm_pred_ix_pb,  lvm_cmr_ix, "Geindexeerd PB-gerecalibreerd"))

            fig2, ax2 = plt.subplots(2, 3, figsize=(18, 11))
            scatter_panel(ax2[0,0], lvm_pred_ix, lvm_cmr_ix, "Geindexeerd ruw",
                          sc_lim2, sc_lim2, color_by=lvh_arr_ix, color_map=cmap)
            ba_panel(ax2[0,1], lvm_pred_ix, lvm_cmr_ix, "Geindexeerd ruw",
                     ba_x_lim2, ba_y_lim2, color_by=lvh_arr_ix, color_map=cmap)
            ba_panel_relative(ax2[0,2], lvm_pred_ix, lvm_cmr_ix, "Geindexeerd ruw",
                              ba_x_lim2, ba_y_lim2_pct, color_by=lvh_arr_ix, color_map=cmap)
            scatter_panel(ax2[1,0], lvm_pred_ix_cal, lvm_cmr_ix, "Geindexeerd OLS-gerecalibreerd",
                          sc_lim2, sc_lim2, color_by=lvh_arr_ix, color_map=cmap)
            ba_panel(ax2[1,1], lvm_pred_ix_cal, lvm_cmr_ix, "Geindexeerd OLS-gerecalibreerd",
                     ba_x_lim2, ba_y_lim2, color_by=lvh_arr_ix, color_map=cmap)
            ba_panel_relative(ax2[1,2], lvm_pred_ix_cal, lvm_cmr_ix, "Geindexeerd OLS-gerecalibreerd",
                              ba_x_lim2, ba_y_lim2_pct, color_by=lvh_arr_ix, color_map=cmap)
            fig2.suptitle(f"[{cohort_name}] LVM-AI vs CMR - Geindexeerde LVM (g/m^2), OLS-kalibratie  (Predicted - CMR)", fontsize=14)
            fig2.tight_layout()
            fig2.savefig(os.path.join(output_dir, "fig2_indexed.png"), dpi=150)
            print(f"  Figuur 2 opgeslagen: {output_dir}/fig2_indexed.png")

    # --- 3. SUBGROEP LVH+ vs LVH- -------------------------------------
    print(f"\n{'='*64}")
    print(f"  3. SUBGROEP: LVH+ vs LVH-")
    print(f"{'='*64}")
    if not has_lvh:
        print("  Geen LVH-kolom - figuur 3 overgeslagen")
    else:
        sub = results.dropna(subset=["lvh_label_int"]).copy()
        sub["lvh_label_int"] = sub["lvh_label_int"].astype(int)
        print(f"  n met geldig label   : {len(sub)} / {len(results)}")
        print(f"  Verdeling            : {sub['lvh_label_int'].value_counts().to_dict()}")

        if sub["lvh_label_int"].nunique() < 2 or len(sub) < 6:
            print("  Te weinig data of slechts een klasse - figuur 3 overgeslagen")
        else:
            g_pos = sub[sub["lvh_label_int"] == 1]
            g_neg = sub[sub["lvh_label_int"] == 0]
            print(f"  LVH+ : n={len(g_pos)},  CMR LVM = {g_pos['lvm_cmr'].mean():.1f} +/- {g_pos['lvm_cmr'].std():.1f} g")
            print(f"  LVH- : n={len(g_neg)},  CMR LVM = {g_neg['lvm_cmr'].mean():.1f} +/- {g_neg['lvm_cmr'].std():.1f} g")

            for grp, df_ in [("LVH+", g_pos), ("LVH-", g_neg)]:
                print(f"\n  -- {grp} --")
                print_metrics(metrics(df_["lvm_predicted"].values,            df_["lvm_cmr"].values, "Absoluut ruw"))
                print_metrics(metrics(df_["lvm_predicted_calibrated"].values, df_["lvm_cmr"].values, "Absoluut OLS-gerecalibreerd"))
                if has_indexed and "lvm_predicted_indexed" in df_.columns:
                    print_metrics(metrics(df_["lvm_predicted_indexed"].values, df_["lvm_cmr_indexed"].values, "Geindexeerd ruw"))
                if has_indexed and "lvm_predicted_indexed_calibrated" in df_.columns:
                    print_metrics(metrics(df_["lvm_predicted_indexed_calibrated"].values,
                                          df_["lvm_cmr_indexed"].values, "Geindexeerd OLS-gerecalibreerd"))

            bias_pos = (g_pos["lvm_predicted"] - g_pos["lvm_cmr"]).values
            bias_neg = (g_neg["lvm_predicted"] - g_neg["lvm_cmr"]).values
            if len(bias_pos) > 1 and len(bias_neg) > 1:
                u, u_p = stats.mannwhitneyu(bias_pos, bias_neg, alternative="two-sided")
                print(f"\n  Mann-Whitney bias LVH+ vs LVH- (Pred - CMR)  U={u:.0f}, p={u_p:.3g}")

            color_by3 = sub["lvh_label_int"].values
            fig3, ax3 = plt.subplots(2, 4, figsize=(22, 10))

            scatter_panel(ax3[0,0], sub["lvm_predicted"].values, sub["lvm_cmr"].values,
                          "Absoluut ruw", sc_lim1, sc_lim1, color_by=color_by3, color_map=LVH_COLOR_MAP)
            scatter_panel(ax3[0,1], sub["lvm_predicted_calibrated"].values, sub["lvm_cmr"].values,
                          "Absoluut OLS-gerecalibreerd", sc_lim1, sc_lim1, color_by=color_by3, color_map=LVH_COLOR_MAP)
            if has_indexed and "lvm_predicted_indexed" in sub.columns and sc_lim2 is not None:
                scatter_panel(ax3[0,2], sub["lvm_predicted_indexed"].values, sub["lvm_cmr_indexed"].values,
                              "Geindexeerd ruw", sc_lim2, sc_lim2, color_by=color_by3, color_map=LVH_COLOR_MAP)
            else:
                ax3[0,2].axis("off")
            if has_indexed and "lvm_predicted_indexed_calibrated" in sub.columns and sc_lim2 is not None:
                scatter_panel(ax3[0,3], sub["lvm_predicted_indexed_calibrated"].values, sub["lvm_cmr_indexed"].values,
                              "Geindexeerd OLS-gerecalibreerd", sc_lim2, sc_lim2, color_by=color_by3, color_map=LVH_COLOR_MAP)
            else:
                ax3[0,3].axis("off")

            for k in range(4):
                ax3[1, k].axis("off")

            fig3.suptitle(f"[{cohort_name}] Subgroup: LVH+ vs LVH- (Spearman)", fontsize=14)
            fig3.tight_layout()
            fig3.savefig(os.path.join(output_dir, "fig3_lvh_subgroep.png"), dpi=150)
            print(f"\n  Figure 3 saved: {output_dir}/fig3_lvh_subgroep.png")

    # --- BOXPLOTS LVH+ vs LVH- ----------------------------------------
    if has_lvh:
        sub_bp = results.dropna(subset=["lvh_label_int"]).copy()
        sub_bp["lvh_label_int"] = sub_bp["lvh_label_int"].astype(int)
        if sub_bp["lvh_label_int"].nunique() >= 2 and len(sub_bp) >= 6:
            bp_pos = sub_bp[sub_bp["lvh_label_int"] == 1]
            bp_neg = sub_bp[sub_bp["lvh_label_int"] == 0]

            panels = [
                ("CMR LV Mass",
                 bp_neg["lvm_cmr"].values,
                 bp_pos["lvm_cmr"].values,
                 "LVM (g)", "abs"),
                ("AI-Predicted LVM\n(Raw)",
                 bp_neg["lvm_predicted"].values,
                 bp_pos["lvm_predicted"].values,
                 "Predicted LVM (g)", "abs"),
                ("AI-Predicted LVM\n(OLS Calibrated)",
                 bp_neg["lvm_predicted_calibrated"].values,
                 bp_pos["lvm_predicted_calibrated"].values,
                 "Predicted LVM (g)", "abs"),
                ("AI-Predicted LVM\n(Passing-Bablok Calibrated)",
                 bp_neg["lvm_predicted_pb"].values,
                 bp_pos["lvm_predicted_pb"].values,
                 "Predicted LVM (g)", "abs"),
            ]

            has_ix_ols = ("lvm_predicted_indexed_calibrated" in sub_bp.columns
                          and bp_pos["lvm_predicted_indexed_calibrated"].notna().any()
                          and bp_neg["lvm_predicted_indexed_calibrated"].notna().any())
            has_ix_pb  = ("lvm_predicted_indexed_pb" in sub_bp.columns
                          and bp_pos["lvm_predicted_indexed_pb"].notna().any()
                          and bp_neg["lvm_predicted_indexed_pb"].notna().any())

            if has_ix_ols:
                panels.append(
                    ("AI-Predicted Indexed LVM\n(OLS Calibrated)",
                     bp_neg["lvm_predicted_indexed_calibrated"].dropna().values,
                     bp_pos["lvm_predicted_indexed_calibrated"].dropna().values,
                     "Predicted LVM (g/m²)", "ix"))
            if has_ix_pb:
                panels.append(
                    ("AI-Predicted Indexed LVM\n(Passing-Bablok Calibrated)",
                     bp_neg["lvm_predicted_indexed_pb"].dropna().values,
                     bp_pos["lvm_predicted_indexed_pb"].dropna().values,
                     "Predicted LVM (g/m²)", "ix"))

            n_panels = len(panels)

            abs_data = []
            ix_data  = []
            for title, d_neg, d_pos, ylabel, group in panels:
                if group == "abs":
                    abs_data.extend([d_neg, d_pos])
                else:
                    ix_data.extend([d_neg, d_pos])

            def shared_lim(arrays, margin=0.08):
                all_vals = np.concatenate([a for a in arrays if len(a) > 0])
                all_vals = all_vals[np.isfinite(all_vals)]
                lo, hi = all_vals.min(), all_vals.max()
                pad = (hi - lo) * margin
                return (lo - pad, hi + pad)

            ylim_abs = shared_lim(abs_data) if abs_data else (0, 1)
            ylim_ix  = shared_lim(ix_data)  if ix_data  else (0, 1)

            if n_panels <= 4:
                fig_bp, axes_bp = plt.subplots(1, n_panels, figsize=(5 * n_panels, 6))
            else:
                n_cols = min(n_panels, 3)
                n_rows = int(np.ceil(n_panels / n_cols))
                fig_bp, axes_bp = plt.subplots(n_rows, n_cols,
                                               figsize=(6 * n_cols, 6 * n_rows))
                axes_bp = axes_bp.ravel()

            if n_panels == 1:
                axes_bp = [axes_bp]

            box_colors = ["steelblue", "crimson"]

            for idx, (title, d_neg, d_pos, ylabel, group) in enumerate(panels):
                ax = axes_bp[idx]
                bp = boxplot_compat(ax, [d_neg, d_pos],
                                    ["LVH−", "LVH+"],
                                    patch_artist=True, widths=0.5)
                for patch, col in zip(bp["boxes"], box_colors):
                    patch.set_facecolor(col)
                    patch.set_alpha(0.6)
                for element in ["whiskers", "caps"]:
                    for line in bp[element]:
                        line.set_color("black")
                for line in bp["medians"]:
                    line.set_color("black")
                    line.set_linewidth(1.5)

                if len(d_neg) > 1 and len(d_pos) > 1:
                    _, p_val = stats.mannwhitneyu(d_neg, d_pos, alternative="two-sided")
                    ax.text(0.05, 0.95, fmt_p_boxplot(p_val),
                            transform=ax.transAxes, va="top", fontsize=10)

                ax.set_title(title, fontsize=11)
                ax.set_ylabel(ylabel)
                ax.set_ylim(ylim_abs if group == "abs" else ylim_ix)

            if hasattr(axes_bp, '__len__'):
                for idx in range(n_panels, len(axes_bp)):
                    axes_bp[idx].axis("off")

            fig_bp.suptitle(
                f"[{cohort_name}] Distribution of LVM Stratified by CMR-Defined Exercise-Induced LVH",
                fontsize=14, y=1.01)
            fig_bp.tight_layout()
            fig_bp.savefig(os.path.join(output_dir, "fig_boxplots_lvh.png"),
                           dpi=150, bbox_inches="tight")
            print(f"\n  Boxplot figure saved: {output_dir}/fig_boxplots_lvh.png")

    # --- 4. PASSING-BABLOK SENSITIVITY --------------------------------
    print(f"\n{'='*64}")
    print(f"  4. SENSITIVITY: Passing-Bablok kalibratie  (Predicted - CMR)")
    print(f"{'='*64}")
    print(f"  Doel: robuustheid van OLS-kalibratie toetsen tegen niet-parametrische PB.")
    print(f"  Absoluut    : CMR = {pb_slope_abs:.3f} * predicted + {pb_intercept_abs:+.1f}")
    if has_indexed and pb_slope_ix is not None:
        print(f"  Geindexeerd : CMR/BSA = {pb_slope_ix:.3f} * predicted/BSA + {pb_intercept_ix:+.1f}")

    print(f"\n  -- Totaal (PB-gerecalibreerd) --")
    print_metrics(metrics(lvm_pred_pb, lvm_cmr, "Absoluut PB"))
    if has_indexed and "lvm_predicted_indexed_pb" in results.columns:
        ix_mask_pb = np.isfinite(results["lvm_predicted_indexed_pb"]) & np.isfinite(results["lvm_cmr_indexed"])
        print_metrics(metrics(results.loc[ix_mask_pb, "lvm_predicted_indexed_pb"].values,
                              results.loc[ix_mask_pb, "lvm_cmr_indexed"].values,
                              "Geindexeerd PB"))

    if has_lvh:
        sub_pb = results.dropna(subset=["lvh_label_int"]).copy()
        sub_pb["lvh_label_int"] = sub_pb["lvh_label_int"].astype(int)
        if sub_pb["lvh_label_int"].nunique() >= 2 and len(sub_pb) >= 6:
            g_pos_pb = sub_pb[sub_pb["lvh_label_int"] == 1]
            g_neg_pb = sub_pb[sub_pb["lvh_label_int"] == 0]

            for grp, df_ in [("LVH+", g_pos_pb), ("LVH-", g_neg_pb)]:
                print(f"\n  -- {grp} (PB-gerecalibreerd) --")
                print_metrics(metrics(df_["lvm_predicted_pb"].values, df_["lvm_cmr"].values, "Absoluut PB"))
                if has_indexed and "lvm_predicted_indexed_pb" in df_.columns:
                    mask_ix = df_["lvm_predicted_indexed_pb"].notna() & df_["lvm_cmr_indexed"].notna()
                    if mask_ix.sum() >= 3:
                        print_metrics(metrics(df_.loc[mask_ix, "lvm_predicted_indexed_pb"].values,
                                              df_.loc[mask_ix, "lvm_cmr_indexed"].values,
                                              "Geindexeerd PB"))

            bias_pos_pb = (g_pos_pb["lvm_predicted_pb"] - g_pos_pb["lvm_cmr"]).values
            bias_neg_pb = (g_neg_pb["lvm_predicted_pb"] - g_neg_pb["lvm_cmr"]).values
            if len(bias_pos_pb) > 1 and len(bias_neg_pb) > 1:
                u_pb, u_p_pb = stats.mannwhitneyu(bias_pos_pb, bias_neg_pb, alternative="two-sided")
                print(f"\n  Mann-Whitney PB-bias LVH+ vs LVH- (Pred - CMR)  U={u_pb:.0f}, p={u_p_pb:.3g}")

            if has_indexed and "lvm_predicted_indexed_pb" in g_pos_pb.columns:
                ix_pos = g_pos_pb.dropna(subset=["lvm_predicted_indexed_pb", "lvm_cmr_indexed"])
                ix_neg = g_neg_pb.dropna(subset=["lvm_predicted_indexed_pb", "lvm_cmr_indexed"])
                if len(ix_pos) > 1 and len(ix_neg) > 1:
                    bias_pos_ix_pb = (ix_pos["lvm_predicted_indexed_pb"] - ix_pos["lvm_cmr_indexed"]).values
                    bias_neg_ix_pb = (ix_neg["lvm_predicted_indexed_pb"] - ix_neg["lvm_cmr_indexed"]).values
                    u_ix_pb, u_p_ix_pb = stats.mannwhitneyu(bias_pos_ix_pb, bias_neg_ix_pb, alternative="two-sided")
                    print(f"  Mann-Whitney PB-bias (geindexeerd) LVH+ vs LVH-  U={u_ix_pb:.0f}, p={u_p_ix_pb:.3g}")

    print(f"\n  -- OLS vs PB vergelijking (totaal) --")
    m_ols_abs = metrics(lvm_pred_cal, lvm_cmr, "OLS abs")
    m_pb_abs  = metrics(lvm_pred_pb,  lvm_cmr, "PB abs")
    print(f"  {'Metric':<20}{'OLS':>15}{'PB':>15}{'Delta(PB-OLS)':>15}")
    print(f"  {'-'*65}")
    print(f"  {'Spearman rho':<20}{m_ols_abs['rho']:>15.3f}{m_pb_abs['rho']:>15.3f}{m_pb_abs['rho']-m_ols_abs['rho']:>+15.3f}")
    print(f"  {'MAE (g)':<20}{m_ols_abs['mae']:>15.1f}{m_pb_abs['mae']:>15.1f}{m_pb_abs['mae']-m_ols_abs['mae']:>+15.1f}")
    print(f"  {'sMAPE (%)':<20}{m_ols_abs['smape']:>15.1f}{m_pb_abs['smape']:>15.1f}{m_pb_abs['smape']-m_ols_abs['smape']:>+15.1f}")
    print(f"  {'Bias (g)':<20}{m_ols_abs['mean_d']:>+15.1f}{m_pb_abs['mean_d']:>+15.1f}{m_pb_abs['mean_d']-m_ols_abs['mean_d']:>+15.1f}")
    print(f"  {'LoA breedte (g)':<20}{m_ols_abs['loa_hi']-m_ols_abs['loa_lo']:>15.1f}{m_pb_abs['loa_hi']-m_pb_abs['loa_lo']:>15.1f}{(m_pb_abs['loa_hi']-m_pb_abs['loa_lo'])-(m_ols_abs['loa_hi']-m_ols_abs['loa_lo']):>+15.1f}")

    if has_indexed and "lvm_predicted_indexed_pb" in results.columns:
        ix_mask_cmp = (np.isfinite(results["lvm_predicted_indexed_pb"]) &
                       np.isfinite(results["lvm_cmr_indexed"]) &
                       np.isfinite(results["lvm_predicted_indexed_calibrated"]))
        if ix_mask_cmp.sum() >= 3:
            m_ols_ix = metrics(results.loc[ix_mask_cmp, "lvm_predicted_indexed_calibrated"].values,
                               results.loc[ix_mask_cmp, "lvm_cmr_indexed"].values, "OLS ix")
            m_pb_ix  = metrics(results.loc[ix_mask_cmp, "lvm_predicted_indexed_pb"].values,
                               results.loc[ix_mask_cmp, "lvm_cmr_indexed"].values, "PB ix")
            print(f"\n  -- OLS vs PB vergelijking (geindexeerd) --")
            print(f"  {'Metric':<20}{'OLS':>15}{'PB':>15}{'Delta(PB-OLS)':>15}")
            print(f"  {'-'*65}")
            print(f"  {'Spearman rho':<20}{m_ols_ix['rho']:>15.3f}{m_pb_ix['rho']:>15.3f}{m_pb_ix['rho']-m_ols_ix['rho']:>+15.3f}")
            print(f"  {'MAE (g/m^2)':<20}{m_ols_ix['mae']:>15.1f}{m_pb_ix['mae']:>15.1f}{m_pb_ix['mae']-m_ols_ix['mae']:>+15.1f}")
            print(f"  {'sMAPE (%)':<20}{m_ols_ix['smape']:>15.1f}{m_pb_ix['smape']:>15.1f}{m_pb_ix['smape']-m_ols_ix['smape']:>+15.1f}")
            print(f"  {'Bias (g/m^2)':<20}{m_ols_ix['mean_d']:>+15.1f}{m_pb_ix['mean_d']:>+15.1f}{m_pb_ix['mean_d']-m_ols_ix['mean_d']:>+15.1f}")
            print(f"  {'LoA breedte (g/m^2)':<20}{m_ols_ix['loa_hi']-m_ols_ix['loa_lo']:>15.1f}{m_pb_ix['loa_hi']-m_pb_ix['loa_lo']:>15.1f}{(m_pb_ix['loa_hi']-m_pb_ix['loa_lo'])-(m_ols_ix['loa_hi']-m_ols_ix['loa_lo']):>+15.1f}")

    fig4, ax4 = plt.subplots(2, 3, figsize=(18, 11))
    scatter_panel(ax4[0,0], lvm_pred_pb, lvm_cmr, "Absoluut PB-gerecalibreerd",
                  sc_lim1, sc_lim1, color_by=lvh_arr, color_map=cmap)
    ba_panel(ax4[0,1], lvm_pred_pb, lvm_cmr, "Absoluut PB-gerecalibreerd",
             ba_x_lim1, ba_y_lim1, color_by=lvh_arr, color_map=cmap)
    ba_panel_relative(ax4[0,2], lvm_pred_pb, lvm_cmr, "Absoluut PB-gerecalibreerd",
                      ba_x_lim1, ba_y_lim1_pct, color_by=lvh_arr, color_map=cmap)

    if has_indexed and "lvm_predicted_indexed_pb" in results.columns and sc_lim2 is not None:
        ix_mask2 = np.isfinite(results["lvm_predicted_indexed_pb"]) & np.isfinite(results["lvm_cmr_indexed"])
        pb_ix_vals  = results.loc[ix_mask2, "lvm_predicted_indexed_pb"].values
        cmr_ix_vals = results.loc[ix_mask2, "lvm_cmr_indexed"].values
        lvh_ix_vals = results.loc[ix_mask2, "lvh_label_int"].values if has_lvh else None
        scatter_panel(ax4[1,0], pb_ix_vals, cmr_ix_vals, "Geindexeerd PB-gerecalibreerd",
                      sc_lim2, sc_lim2, color_by=lvh_ix_vals, color_map=cmap)
        ba_panel(ax4[1,1], pb_ix_vals, cmr_ix_vals, "Geindexeerd PB-gerecalibreerd",
                 ba_x_lim2, ba_y_lim2, color_by=lvh_ix_vals, color_map=cmap)
        ba_panel_relative(ax4[1,2], pb_ix_vals, cmr_ix_vals, "Geindexeerd PB-gerecalibreerd",
                          ba_x_lim2, ba_y_lim2_pct, color_by=lvh_ix_vals, color_map=cmap)
    else:
        for k in range(3): ax4[1,k].axis("off")

    fig4.suptitle(f"[{cohort_name}] Sensitivity analysis - Passing-Bablok kalibratie  (Predicted - CMR)", fontsize=14)
    fig4.tight_layout()
    fig4.savefig(os.path.join(output_dir, "fig4_passing_bablok.png"), dpi=150)
    print(f"\n  Figuur 4 opgeslagen: {output_dir}/fig4_passing_bablok.png")

    # --- CSV met voorspellingen voor deze cohort ----------------------
    output_csv = os.path.join(output_dir, "regressie_voorspellingen.csv")
    results.to_csv(output_csv, index=False)
    print(f"\nVoorspellingen opgeslagen: {output_csv}")

    # --- 5. SCATTER PANEL (2x2) ---------------------------------------
    print(f"\n{'='*64}")
    print(f"  5. COMBINED SCATTER FIGURE (2x2)  -  English labels")
    print(f"{'='*64}")

    fig5, ax5 = plt.subplots(2, 2, figsize=(14, 12))

    scatter_panel_en(
        ax5[0, 0], lvm_pred_cal, lvm_cmr,
        xlabel="LVM-AI LVM (g)", ylabel="CMR-derived LVM (g)",
        panel_letter="A", panel_title="Linear Recalibrated",
        xlim=sc_lim1, ylim=sc_lim1,
        color_by=lvh_arr, color_map=cmap,
    )

    if has_indexed and sc_lim2 is not None and "lvm_predicted_indexed_calibrated" in results.columns:
        ix_mask_b = (np.isfinite(results["lvm_predicted_indexed_calibrated"]) &
                     np.isfinite(results["lvm_cmr_indexed"]))
        pred_b = results.loc[ix_mask_b, "lvm_predicted_indexed_calibrated"].values
        cmr_b  = results.loc[ix_mask_b, "lvm_cmr_indexed"].values
        lvh_b  = results.loc[ix_mask_b, "lvh_label_int"].values if has_lvh else None
        scatter_panel_en(
            ax5[0, 1], pred_b, cmr_b,
            xlabel="LVM-AI indexed LVM (g/m²)",
            ylabel="CMR-derived indexed LVM (g/m²)",
            panel_letter="B", panel_title="Indexed Linear Recalibrated",
            xlim=sc_lim2, ylim=sc_lim2,
            color_by=lvh_b, color_map=cmap,
        )
    else:
        ax5[0, 1].axis("off")

    scatter_panel_en(
        ax5[1, 0], lvm_pred_pb, lvm_cmr,
        xlabel="LVM-AI LVM (g)", ylabel="CMR-derived LVM (g)",
        panel_letter="C", panel_title="Passing-Bablok Recalibrated",
        xlim=sc_lim1, ylim=sc_lim1,
        color_by=lvh_arr, color_map=cmap,
    )

    if has_indexed and sc_lim2 is not None and "lvm_predicted_indexed_pb" in results.columns:
        ix_mask_d = (np.isfinite(results["lvm_predicted_indexed_pb"]) &
                     np.isfinite(results["lvm_cmr_indexed"]))
        pred_d = results.loc[ix_mask_d, "lvm_predicted_indexed_pb"].values
        cmr_d  = results.loc[ix_mask_d, "lvm_cmr_indexed"].values
        lvh_d  = results.loc[ix_mask_d, "lvh_label_int"].values if has_lvh else None
        scatter_panel_en(
            ax5[1, 1], pred_d, cmr_d,
            xlabel="LVM-AI indexed LVM (g/m²)",
            ylabel="CMR-derived indexed LVM (g/m²)",
            panel_letter="D", panel_title="Indexed Passing-Bablok Recalibrated",
            xlim=sc_lim2, ylim=sc_lim2,
            color_by=lvh_d, color_map=cmap,
        )
    else:
        ax5[1, 1].axis("off")

    fig5.suptitle(f"[{cohort_name}] LVM-AI vs CMR-derived LVM  -  Spearman Correlations",
                  fontsize=14, y=0.995)
    fig5.tight_layout()
    fig5.savefig(os.path.join(output_dir, "fig5_scatter_2x2.png"), dpi=150)
    print(f"  Figure 5 saved: {output_dir}/fig5_scatter_2x2.png")

    # --- 6. RELATIVE BA (indexed): OLS vs PB --------------------------
    if (has_indexed and sc_lim2 is not None
            and "lvm_predicted_indexed_calibrated" in results.columns
            and "lvm_predicted_indexed_pb" in results.columns):

        print(f"\n{'='*64}")
        print(f"  6. RELATIVE BA (indexed): OLS vs PB - zero-centered shared axis")
        print(f"{'='*64}")

        ix_mask_ba = (np.isfinite(results["lvm_predicted_indexed_calibrated"]) &
                      np.isfinite(results["lvm_predicted_indexed_pb"]) &
                      np.isfinite(results["lvm_cmr_indexed"]))
        pred_ols_ix = results.loc[ix_mask_ba, "lvm_predicted_indexed_calibrated"].values
        pred_pb_ix  = results.loc[ix_mask_ba, "lvm_predicted_indexed_pb"].values
        cmr_ix_ba   = results.loc[ix_mask_ba, "lvm_cmr_indexed"].values
        lvh_ix_ba   = results.loc[ix_mask_ba, "lvh_label_int"].values if has_lvh else None

        pct_ols = pct_diffs(pred_ols_ix, cmr_ix_ba)
        pct_pb  = pct_diffs(pred_pb_ix,  cmr_ix_ba)
        all_pct = np.concatenate([pct_ols, pct_pb])
        md_ols, sd_ols = pct_ols.mean(), pct_ols.std()
        md_pb,  sd_pb  = pct_pb.mean(),  pct_pb.std()
        abs_extreme = max(
            abs(md_ols + 1.96*sd_ols), abs(md_ols - 1.96*sd_ols),
            abs(md_pb  + 1.96*sd_pb),  abs(md_pb  - 1.96*sd_pb),
            float(np.nanmax(np.abs(all_pct))),
        )
        ylim_sym_pct = (-abs_extreme * 1.10, abs_extreme * 1.10)

        mean_all = np.concatenate([(pred_ols_ix + cmr_ix_ba) / 2,
                                   (pred_pb_ix  + cmr_ix_ba) / 2])
        xlim_ba6 = data_lims(mean_all)

        fig6, ax6 = plt.subplots(1, 2, figsize=(16, 7))
        ba_panel_centered(
            ax6[0], pred_ols_ix, cmr_ix_ba,
            panel_letter="A",
            title_text="OLS-recalibrated, indexed LVM",
            xlim=xlim_ba6, ylim_sym=ylim_sym_pct,
            color_by=lvh_ix_ba, color_map=cmap, relative=True,
        )
        ba_panel_centered(
            ax6[1], pred_pb_ix, cmr_ix_ba,
            panel_letter="B",
            title_text="Passing-Bablok-recalibrated, indexed LVM",
            xlim=xlim_ba6, ylim_sym=ylim_sym_pct,
            color_by=lvh_ix_ba, color_map=cmap, relative=True,
        )
        for a in ax6:
            a.set_xlabel("Mean of CMR and predicted indexed LVM (g/m²)")

        fig6.suptitle(
            f"[{cohort_name}] Relative Bland-Altman - indexed LVM (g/m²)  (Predicted - CMR)",
            fontsize=14, y=0.99,
        )
        foot6 = (
            f"OLS (A): CMR/BSA = {coef_pred_ix:.3f} × predicted/BSA "
            f"{coef_sex_ix:+.2f} × sex {intercept_ix:+.2f}    |    "
            f"Passing-Bablok (B): CMR/BSA = {pb_slope_ix:.3f} × predicted/BSA "
            f"{pb_intercept_ix:+.2f}  (univariate). "
            f"Positive = AI overestimates; negative = AI underestimates."
        )
        fig6.text(0.5, 0.01, foot6, ha="center", va="bottom", fontsize=9,
                  style="italic")
        fig6.tight_layout(rect=[0, 0.04, 1, 0.97])
        fig6.savefig(os.path.join(output_dir, "fig6_BA_relative_indexed.png"), dpi=150)
        print(f"  Figure 6 saved: {output_dir}/fig6_BA_relative_indexed.png")
    else:
        print("\n  Figure 6 skipped (indexed data or PB calibration missing)")

    # --- 7. ABSOLUTE BA (g): OLS vs PB --------------------------------
    print(f"\n{'='*64}")
    print(f"  7. ABSOLUTE BA (g): OLS vs PB - zero-centered shared axis")
    print(f"{'='*64}")

    diff_ols_abs = lvm_pred_cal - lvm_cmr
    diff_pb_abs  = lvm_pred_pb  - lvm_cmr
    md_ols_a, sd_ols_a = diff_ols_abs.mean(), diff_ols_abs.std()
    md_pb_a,  sd_pb_a  = diff_pb_abs.mean(),  diff_pb_abs.std()
    abs_extreme_g = max(
        abs(md_ols_a + 1.96*sd_ols_a), abs(md_ols_a - 1.96*sd_ols_a),
        abs(md_pb_a  + 1.96*sd_pb_a),  abs(md_pb_a  - 1.96*sd_pb_a),
        float(np.nanmax(np.abs(np.concatenate([diff_ols_abs, diff_pb_abs])))),
    )
    ylim_sym_g = (-abs_extreme_g * 1.10, abs_extreme_g * 1.10)

    mean_all_g = np.concatenate([(lvm_pred_cal + lvm_cmr) / 2,
                                 (lvm_pred_pb  + lvm_cmr) / 2])
    xlim_ba7 = data_lims(mean_all_g)

    fig7, ax7 = plt.subplots(1, 2, figsize=(16, 7))
    ba_panel_centered(
        ax7[0], lvm_pred_cal, lvm_cmr,
        panel_letter="A",
        title_text="OLS-recalibrated, absolute LVM",
        xlim=xlim_ba7, ylim_sym=ylim_sym_g,
        color_by=lvh_arr, color_map=cmap, relative=False,
    )
    ba_panel_centered(
        ax7[1], lvm_pred_pb, lvm_cmr,
        panel_letter="B",
        title_text="Passing-Bablok-recalibrated, absolute LVM",
        xlim=xlim_ba7, ylim_sym=ylim_sym_g,
        color_by=lvh_arr, color_map=cmap, relative=False,
    )
    for a in ax7:
        a.set_xlabel("Mean of CMR and predicted LVM (g)")
        a.set_ylabel("Predicted - CMR LVM (g)")

    fig7.suptitle(
        f"[{cohort_name}] Absolute Bland-Altman - LVM (g)  (Predicted - CMR)",
        fontsize=14, y=0.99,
    )
    foot7 = (
        f"OLS (A): CMR = {coef_pred_abs:.3f} × predicted "
        f"{coef_sex_abs:+.2f} × sex {intercept_abs:+.2f}    |    "
        f"Passing-Bablok (B): CMR = {pb_slope_abs:.3f} × predicted "
        f"{pb_intercept_abs:+.2f}  (univariate). "
        f"Positive = AI overestimates; negative = AI underestimates."
    )
    fig7.text(0.5, 0.01, foot7, ha="center", va="bottom", fontsize=9, style="italic")
    fig7.tight_layout(rect=[0, 0.04, 1, 0.97])
    fig7.savefig(os.path.join(output_dir, "fig7_BA_absolute.png"), dpi=150)
    print(f"  Figure 7 saved: {output_dir}/fig7_BA_absolute.png")

    # --- 8. RELATIVE BA (1x3) -----------------------------------------
    print(f"\n{'='*64}")
    print(f"  8. RELATIVE BA (1x3): raw / OLS indexed / PB indexed")
    print(f"{'='*64}")

    pred_A = lvm_pred
    truth_A = lvm_cmr
    lvh_A = lvh_arr
    pct_A = pct_diffs(pred_A, truth_A)

    has_B = (has_indexed and "lvm_predicted_indexed_calibrated" in results.columns)
    has_C = (has_indexed and "lvm_predicted_indexed_pb" in results.columns)

    pct_B = pct_C = None
    pred_B = pred_C = truth_BC = lvh_BC = None
    if has_B and has_C:
        ix_mask_8 = (np.isfinite(results["lvm_predicted_indexed_calibrated"]) &
                     np.isfinite(results["lvm_predicted_indexed_pb"]) &
                     np.isfinite(results["lvm_cmr_indexed"]))
        pred_B = results.loc[ix_mask_8, "lvm_predicted_indexed_calibrated"].values
        pred_C = results.loc[ix_mask_8, "lvm_predicted_indexed_pb"].values
        truth_BC = results.loc[ix_mask_8, "lvm_cmr_indexed"].values
        lvh_BC = results.loc[ix_mask_8, "lvh_label_int"].values if has_lvh else None
        pct_B = pct_diffs(pred_B, truth_BC)
        pct_C = pct_diffs(pred_C, truth_BC)

    all_pct_parts = [pct_A]
    if pct_B is not None: all_pct_parts.append(pct_B)
    if pct_C is not None: all_pct_parts.append(pct_C)
    all_pct = np.concatenate(all_pct_parts)

    def _loa_extent(pct):
        md, sd = pct.mean(), pct.std()
        return max(abs(md + 1.96*sd), abs(md - 1.96*sd))

    loa_max = max(_loa_extent(p) for p in all_pct_parts)
    data_max = float(np.nanmax(np.abs(all_pct)))
    abs_extreme = max(loa_max, data_max)
    ylim_sym_8 = (-abs_extreme * 1.10, abs_extreme * 1.10)

    fig8, ax8 = plt.subplots(1, 3, figsize=(21, 7))

    xlim_A = data_lims((pred_A + truth_A) / 2.0)
    stats_A = ba_relative_strict(
        ax8[0], pred_A, truth_A,
        panel_letter="A",
        subplot_title="Raw Uncalibrated LVM (g)",
        xlabel="Mean LVM (g)",
        xlim=xlim_A, ylim_sym=ylim_sym_8,
        color_by=lvh_A, color_map=cmap,
    )

    if pct_B is not None and pct_C is not None:
        mean_BC = np.concatenate([(pred_B + truth_BC) / 2.0,
                                  (pred_C + truth_BC) / 2.0])
        xlim_BC = data_lims(mean_BC)

        stats_B = ba_relative_strict(
            ax8[1], pred_B, truth_BC,
            panel_letter="B",
            subplot_title="Indexed Calibrated OLS LVM (g/m²)",
            xlabel="Mean Indexed LVM (g/m²)",
            xlim=xlim_BC, ylim_sym=ylim_sym_8,
            color_by=lvh_BC, color_map=cmap,
        )
        stats_C = ba_relative_strict(
            ax8[2], pred_C, truth_BC,
            panel_letter="C",
            subplot_title="Indexed Calibrated Passing-Bablok LVM (g/m²)",
            xlabel="Mean Indexed LVM (g/m²)",
            xlim=xlim_BC, ylim_sym=ylim_sym_8,
            color_by=lvh_BC, color_map=cmap,
        )
    else:
        for k in (1, 2):
            ax8[k].axis("off")
            ax8[k].set_title("Indexed data not available")
        stats_B = stats_C = None

    fig8.suptitle(f"[{cohort_name}] LVM-AI vs CMR-derived LVM - Bland-Altman Plots  (Predicted - CMR)",
                  fontsize=14, y=0.995)
    fig8.text(0.5, 0.005,
              "Positive values = AI overestimates LVM; negative values = AI underestimates LVM.",
              ha="center", va="bottom", fontsize=9, style="italic")
    fig8.tight_layout(rect=[0, 0.03, 1, 0.97])
    fig8.savefig(os.path.join(output_dir, "fig8_BA_relative_1x3.png"), dpi=150)
    print(f"  Figure 8 saved: {output_dir}/fig8_BA_relative_1x3.png")

    for letter, s in [("A", stats_A), ("B", stats_B), ("C", stats_C)]:
        if s is None: continue
        print(f"  ({letter})  n={s['n']:3d}  Bias={s['bias']:+.1f}%  "
              f"Upper LoA={s['upper']:+.1f}%  Lower LoA={s['lower']:+.1f}%")

    # --- 9. EXCEL WORKBOOK --------------------------------------------
    print(f"\n{'='*64}")
    print(f"  9. EXCEL WORKBOOK (6 sheets: absolute + relative)")
    print(f"{'='*64}")

    def _M(pred, truth, name):
        pred = np.asarray(pred, dtype=float); truth = np.asarray(truth, dtype=float)
        mask = np.isfinite(pred) & np.isfinite(truth)
        if mask.sum() < 3:
            return None
        return metrics(pred[mask], truth[mask], name)

    M_TOT = {
        "uncal_abs": _M(results["lvm_predicted"].values,
                        results["lvm_cmr"].values, "tot_uncal_abs"),
        "ols_abs":   _M(results["lvm_predicted_calibrated"].values,
                        results["lvm_cmr"].values, "tot_ols_abs"),
        "pb_abs":    _M(results["lvm_predicted_pb"].values,
                        results["lvm_cmr"].values, "tot_pb_abs"),
        "uncal_ix":  None, "ols_ix": None, "pb_ix": None,
    }
    if has_indexed:
        if "lvm_predicted_indexed" in results.columns:
            M_TOT["uncal_ix"] = _M(results["lvm_predicted_indexed"].values,
                                   results["lvm_cmr_indexed"].values, "tot_uncal_ix")
        if "lvm_predicted_indexed_calibrated" in results.columns:
            M_TOT["ols_ix"]   = _M(results["lvm_predicted_indexed_calibrated"].values,
                                   results["lvm_cmr_indexed"].values, "tot_ols_ix")
        if "lvm_predicted_indexed_pb" in results.columns:
            M_TOT["pb_ix"]    = _M(results["lvm_predicted_indexed_pb"].values,
                                   results["lvm_cmr_indexed"].values, "tot_pb_ix")

    def _group_metrics(df_):
        out = {
            "uncal_abs": _M(df_["lvm_predicted"].values, df_["lvm_cmr"].values, "uncal_abs"),
            "ols_abs":   _M(df_["lvm_predicted_calibrated"].values, df_["lvm_cmr"].values, "ols_abs"),
            "pb_abs":    _M(df_["lvm_predicted_pb"].values, df_["lvm_cmr"].values, "pb_abs"),
            "uncal_ix": None, "ols_ix": None, "pb_ix": None,
        }
        if has_indexed:
            if "lvm_predicted_indexed" in df_.columns:
                out["uncal_ix"] = _M(df_["lvm_predicted_indexed"].values,
                                     df_["lvm_cmr_indexed"].values, "uncal_ix")
            if "lvm_predicted_indexed_calibrated" in df_.columns:
                out["ols_ix"]   = _M(df_["lvm_predicted_indexed_calibrated"].values,
                                     df_["lvm_cmr_indexed"].values, "ols_ix")
            if "lvm_predicted_indexed_pb" in df_.columns:
                out["pb_ix"]    = _M(df_["lvm_predicted_indexed_pb"].values,
                                     df_["lvm_cmr_indexed"].values, "pb_ix")
        return out

    M_POS = M_NEG = None
    n_pos = n_neg = 0
    mw = {"uncal_abs": np.nan, "ols_abs": np.nan, "pb_abs": np.nan,
          "uncal_ix": np.nan, "ols_ix": np.nan, "pb_ix": np.nan}

    if has_lvh:
        sub_x = results.dropna(subset=["lvh_label_int"]).copy()
        sub_x["lvh_label_int"] = sub_x["lvh_label_int"].astype(int)
        g_pos_x = sub_x[sub_x["lvh_label_int"] == 1]
        g_neg_x = sub_x[sub_x["lvh_label_int"] == 0]
        n_pos, n_neg = len(g_pos_x), len(g_neg_x)
        if n_pos > 0 and n_neg > 0:
            M_POS = _group_metrics(g_pos_x)
            M_NEG = _group_metrics(g_neg_x)

            def _mw(pos_pred_col, neg_pred_col, pos_truth_col, neg_truth_col):
                d_pos = (g_pos_x[pos_pred_col] - g_pos_x[pos_truth_col]).dropna().values
                d_neg = (g_neg_x[neg_pred_col] - g_neg_x[neg_truth_col]).dropna().values
                if len(d_pos) > 1 and len(d_neg) > 1:
                    _, p = stats.mannwhitneyu(d_pos, d_neg, alternative="two-sided")
                    return p
                return np.nan

            mw["uncal_abs"] = _mw("lvm_predicted",            "lvm_predicted",
                                  "lvm_cmr",                   "lvm_cmr")
            mw["ols_abs"]   = _mw("lvm_predicted_calibrated", "lvm_predicted_calibrated",
                                  "lvm_cmr",                   "lvm_cmr")
            mw["pb_abs"]    = _mw("lvm_predicted_pb",         "lvm_predicted_pb",
                                  "lvm_cmr",                   "lvm_cmr")
            if has_indexed:
                if "lvm_predicted_indexed" in g_pos_x.columns:
                    mw["uncal_ix"] = _mw("lvm_predicted_indexed", "lvm_predicted_indexed",
                                         "lvm_cmr_indexed",       "lvm_cmr_indexed")
                if "lvm_predicted_indexed_calibrated" in g_pos_x.columns:
                    mw["ols_ix"]   = _mw("lvm_predicted_indexed_calibrated",
                                         "lvm_predicted_indexed_calibrated",
                                         "lvm_cmr_indexed", "lvm_cmr_indexed")
                if "lvm_predicted_indexed_pb" in g_pos_x.columns:
                    mw["pb_ix"]    = _mw("lvm_predicted_indexed_pb",
                                         "lvm_predicted_indexed_pb",
                                         "lvm_cmr_indexed", "lvm_cmr_indexed")

    SIGN_NOTE = ("Sign convention: Bias = Predicted - CMR. Positive values "
                 "indicate AI overestimation; negative values indicate AI underestimation.")

    ABBR_ABS = ("Abbreviations: CI, confidence interval; CMR, cardiovascular magnetic "
                "resonance; LoA, limits of agreement; LVM, left ventricular mass; "
                "LVMi, indexed left ventricular mass; MAE, mean absolute error; "
                "OLS, ordinary least squares.")
    ABBR_REL = ("Abbreviations: CI, confidence interval; CMR, cardiovascular magnetic "
                "resonance; LoA, limits of agreement; LVM, left ventricular mass; "
                "LVMi, indexed left ventricular mass; OLS, ordinary least squares; "
                "sMAPE, symmetric mean absolute percentage error. All percentage values "
                "expressed relative to the mean of CMR and predicted (Bland-Altman).")

    FORMULA_OLS_ABS = (f"OLS calibration (absolute): CMR = {coef_pred_abs:.3f} × predicted "
                       f"{coef_sex_abs:+.2f} × sex {intercept_abs:+.2f}.")
    FORMULA_PB_ABS  = (f"Passing-Bablok (absolute): CMR = {pb_slope_abs:.3f} × predicted "
                       f"{pb_intercept_abs:+.2f} (univariate).")
    FORMULA_OLS_IX = FORMULA_PB_IX = ""
    if has_indexed and coef_pred_ix is not None:
        FORMULA_OLS_IX = (f"OLS calibration (indexed): CMR/BSA = {coef_pred_ix:.3f} × predicted/BSA "
                          f"{coef_sex_ix:+.2f} × sex {intercept_ix:+.2f}.")
    if has_indexed and pb_slope_ix is not None:
        FORMULA_PB_IX  = (f"Passing-Bablok (indexed): CMR/BSA = {pb_slope_ix:.3f} × predicted/BSA "
                          f"{pb_intercept_ix:+.2f} (univariate).")

    wb = Workbook()
    ws = wb.active; ws.title = "Abs-Total"

    n_total = len(results)
    COLS_ABS_TOT = ["Uncalibrated", "Linear (OLS)", "Passing-Bablok"]

    rows_a1 = []
    rows_a1.append(("Absolute LVM (g)", ["", "", ""], True))
    rows_a1.append(("Reference CMR LVM (g)",
                    [f_ref(results["lvm_cmr"].values), "", ""], False))
    rows_a1.append(("Predicted LVM (g)",
                    [f_ref(results["lvm_predicted"].values),
                     f_ref(results["lvm_predicted_calibrated"].values),
                     f_ref(results["lvm_predicted_pb"].values)], False))
    rows_a1.append(("Spearman ρ (95% CI)",
                    [f_rho(M_TOT["uncal_abs"]), f_rho(M_TOT["ols_abs"]),
                     f_rho(M_TOT["pb_abs"])], False))
    rows_a1.append(("MAE, g (95% CI)",
                    [f_mae(M_TOT["uncal_abs"]), f_mae(M_TOT["ols_abs"]),
                     f_mae(M_TOT["pb_abs"])], False))
    rows_a1.append(("Mean Bias, g (Predicted - CMR)",
                    [f_bias_abs(M_TOT["uncal_abs"]), f_bias_abs(M_TOT["ols_abs"]),
                     f_bias_abs(M_TOT["pb_abs"])], False))
    rows_a1.append(("95% LoA, g",
                    [f_loa_abs(M_TOT["uncal_abs"]), f_loa_abs(M_TOT["ols_abs"]),
                     f_loa_abs(M_TOT["pb_abs"])], False))

    if has_indexed:
        rows_a1.append(("Indexed LVMi (g/m²)", ["", "", ""], True))
        rows_a1.append(("Reference CMR LVMi (g/m²)",
                        [f_ref(results["lvm_cmr_indexed"].dropna().values), "", ""], False))
        rows_a1.append(("Predicted LVMi (g/m²)",
                        [f_ref(results["lvm_predicted_indexed"].dropna().values)
                           if "lvm_predicted_indexed" in results.columns else "-",
                         f_ref(results["lvm_predicted_indexed_calibrated"].dropna().values)
                           if "lvm_predicted_indexed_calibrated" in results.columns else "-",
                         f_ref(results["lvm_predicted_indexed_pb"].dropna().values)
                           if "lvm_predicted_indexed_pb" in results.columns else "-"], False))
        rows_a1.append(("Spearman ρ (95% CI)",
                        [f_rho(M_TOT["uncal_ix"]), f_rho(M_TOT["ols_ix"]),
                         f_rho(M_TOT["pb_ix"])], False))
        rows_a1.append(("MAE, g/m² (95% CI)",
                        [f_mae(M_TOT["uncal_ix"]), f_mae(M_TOT["ols_ix"]),
                         f_mae(M_TOT["pb_ix"])], False))
        rows_a1.append(("Mean Bias, g/m² (Predicted - CMR)",
                        [f_bias_abs(M_TOT["uncal_ix"]), f_bias_abs(M_TOT["ols_ix"]),
                         f_bias_abs(M_TOT["pb_ix"])], False))
        rows_a1.append(("95% LoA, g/m²",
                        [f_loa_abs(M_TOT["uncal_ix"]), f_loa_abs(M_TOT["ols_ix"]),
                         f_loa_abs(M_TOT["pb_ix"])], False))

    foots_a1 = [SIGN_NOTE, FORMULA_OLS_ABS + " " + FORMULA_PB_ABS]
    if FORMULA_OLS_IX: foots_a1.append(FORMULA_OLS_IX + " " + FORMULA_PB_IX)
    foots_a1.append(ABBR_ABS)

    _write_sheet(
        ws, title=f"Table A1. Absolute performance in the {cohort_name} cohort (n = {n_total})",
        col_headers=COLS_ABS_TOT, rows=rows_a1, footnotes=foots_a1,
        col_widths=[40, 26, 26, 26],
    )

    if has_lvh and M_POS is not None and M_NEG is not None:
        ws = wb.create_sheet("Abs-LVH-OLS")
        COLS_LVH = [f"Exercise-Induced LVH\n(n = {n_pos})",
                    f"No Exercise-Induced LVH\n(n = {n_neg})"]
        rows_a2 = []
        rows_a2.append(("Uncalibrated absolute LVM (g)", ["", ""], True))
        rows_a2.append(("Reference CMR LVM (g)",
                        [f_ref(g_pos_x["lvm_cmr"].values),
                         f_ref(g_neg_x["lvm_cmr"].values)], False))
        rows_a2.append(("Spearman ρ (95% CI)",
                        [f_rho(M_POS["uncal_abs"]), f_rho(M_NEG["uncal_abs"])], False))
        rows_a2.append(("MAE, g (95% CI)",
                        [f_mae(M_POS["uncal_abs"]), f_mae(M_NEG["uncal_abs"])], False))
        rows_a2.append(("Mean Bias, g (Predicted - CMR)*",
                        [f_bias_abs(M_POS["uncal_abs"]), f_bias_abs(M_NEG["uncal_abs"])], False))
        rows_a2.append(("95% LoA, g",
                        [f_loa_abs(M_POS["uncal_abs"]), f_loa_abs(M_NEG["uncal_abs"])], False))

        rows_a2.append(("Linear recalibrated absolute LVM (g)", ["", ""], True))
        rows_a2.append(("Spearman ρ (95% CI)",
                        [f_rho(M_POS["ols_abs"]), f_rho(M_NEG["ols_abs"])], False))
        rows_a2.append(("MAE, g (95% CI)",
                        [f_mae(M_POS["ols_abs"]), f_mae(M_NEG["ols_abs"])], False))
        rows_a2.append(("Mean Bias, g (Predicted - CMR)†",
                        [f_bias_abs(M_POS["ols_abs"]), f_bias_abs(M_NEG["ols_abs"])], False))
        rows_a2.append(("95% LoA, g",
                        [f_loa_abs(M_POS["ols_abs"]), f_loa_abs(M_NEG["ols_abs"])], False))

        if has_indexed:
            rows_a2.append(("Linear recalibrated indexed LVMi (g/m²)", ["", ""], True))
            rows_a2.append(("Reference CMR LVMi (g/m²)",
                            [f_ref(g_pos_x["lvm_cmr_indexed"].dropna().values),
                             f_ref(g_neg_x["lvm_cmr_indexed"].dropna().values)], False))
            rows_a2.append(("Spearman ρ (95% CI)",
                            [f_rho(M_POS["ols_ix"]), f_rho(M_NEG["ols_ix"])], False))
            rows_a2.append(("MAE, g/m² (95% CI)",
                            [f_mae(M_POS["ols_ix"]), f_mae(M_NEG["ols_ix"])], False))
            rows_a2.append(("Mean Bias, g/m² (Predicted - CMR)‡",
                            [f_bias_abs(M_POS["ols_ix"]), f_bias_abs(M_NEG["ols_ix"])], False))
            rows_a2.append(("95% LoA, g/m²",
                            [f_loa_abs(M_POS["ols_ix"]), f_loa_abs(M_NEG["ols_ix"])], False))

        foots_a2 = [
            SIGN_NOTE,
            f"* Bias LVH+ vs LVH- (uncalibrated absolute): Mann-Whitney U, {f_p(mw['uncal_abs'])}.",
            f"† Bias LVH+ vs LVH- (OLS recalibrated absolute): Mann-Whitney U, {f_p(mw['ols_abs'])}.",
        ]
        if has_indexed:
            foots_a2.append(
                f"‡ Bias LVH+ vs LVH- (OLS recalibrated indexed): Mann-Whitney U, {f_p(mw['ols_ix'])}.")
        foots_a2.append(ABBR_ABS)

        _write_sheet(
            ws,
            title="Table A2. Absolute performance stratified by exercise-induced LVH (linear OLS recalibration)",
            col_headers=COLS_LVH, rows=rows_a2, footnotes=foots_a2,
            col_widths=[42, 28, 28],
        )

        ws = wb.create_sheet("Abs-LVH-PB")
        rows_a3 = []
        rows_a3.append(("Passing-Bablok recalibrated absolute LVM (g)", ["", ""], True))
        rows_a3.append(("Reference CMR LVM (g)",
                        [f_ref(g_pos_x["lvm_cmr"].values),
                         f_ref(g_neg_x["lvm_cmr"].values)], False))
        rows_a3.append(("Spearman ρ (95% CI)",
                        [f_rho(M_POS["pb_abs"]), f_rho(M_NEG["pb_abs"])], False))
        rows_a3.append(("MAE, g (95% CI)",
                        [f_mae(M_POS["pb_abs"]), f_mae(M_NEG["pb_abs"])], False))
        rows_a3.append(("Mean Bias, g (Predicted - CMR)*",
                        [f_bias_abs(M_POS["pb_abs"]), f_bias_abs(M_NEG["pb_abs"])], False))
        rows_a3.append(("95% LoA, g",
                        [f_loa_abs(M_POS["pb_abs"]), f_loa_abs(M_NEG["pb_abs"])], False))

        if has_indexed:
            rows_a3.append(("Passing-Bablok recalibrated indexed LVMi (g/m²)", ["", ""], True))
            rows_a3.append(("Reference CMR LVMi (g/m²)",
                            [f_ref(g_pos_x["lvm_cmr_indexed"].dropna().values),
                             f_ref(g_neg_x["lvm_cmr_indexed"].dropna().values)], False))
            rows_a3.append(("Spearman ρ (95% CI)",
                            [f_rho(M_POS["pb_ix"]), f_rho(M_NEG["pb_ix"])], False))
            rows_a3.append(("MAE, g/m² (95% CI)",
                            [f_mae(M_POS["pb_ix"]), f_mae(M_NEG["pb_ix"])], False))
            rows_a3.append(("Mean Bias, g/m² (Predicted - CMR)†",
                            [f_bias_abs(M_POS["pb_ix"]), f_bias_abs(M_NEG["pb_ix"])], False))
            rows_a3.append(("95% LoA, g/m²",
                            [f_loa_abs(M_POS["pb_ix"]), f_loa_abs(M_NEG["pb_ix"])], False))

        foots_a3 = [
            SIGN_NOTE,
            f"* Bias LVH+ vs LVH- (PB recalibrated absolute): Mann-Whitney U, {f_p(mw['pb_abs'])}.",
        ]
        if has_indexed:
            foots_a3.append(
                f"† Bias LVH+ vs LVH- (PB recalibrated indexed): Mann-Whitney U, {f_p(mw['pb_ix'])}.")
        foots_a3.append("Sensitivity analysis: Passing-Bablok is non-parametric and does "
                        "not include sex as a covariate.")
        foots_a3.append(ABBR_ABS)

        _write_sheet(
            ws,
            title="Table A3. Absolute performance stratified by exercise-induced LVH (Passing-Bablok, sensitivity)",
            col_headers=COLS_LVH, rows=rows_a3, footnotes=foots_a3,
            col_widths=[42, 28, 28],
        )

    ws = wb.create_sheet("Rel-Total")
    rows_r1 = []
    rows_r1.append(("Relative agreement - absolute LVM (%)", ["", "", ""], True))
    rows_r1.append(("Spearman ρ (95% CI)",
                    [f_rho(M_TOT["uncal_abs"]), f_rho(M_TOT["ols_abs"]),
                     f_rho(M_TOT["pb_abs"])], False))
    rows_r1.append(("sMAPE (%)",
                    [f_smape(M_TOT["uncal_abs"]), f_smape(M_TOT["ols_abs"]),
                     f_smape(M_TOT["pb_abs"])], False))
    rows_r1.append(("Mean Bias (%)  (Predicted - CMR)",
                    [f_bias_pct(M_TOT["uncal_abs"]), f_bias_pct(M_TOT["ols_abs"]),
                     f_bias_pct(M_TOT["pb_abs"])], False))
    rows_r1.append(("95% LoA (%)",
                    [f_loa_pct(M_TOT["uncal_abs"]), f_loa_pct(M_TOT["ols_abs"]),
                     f_loa_pct(M_TOT["pb_abs"])], False))

    if has_indexed:
        rows_r1.append(("Relative agreement - indexed LVMi (%)", ["", "", ""], True))
        rows_r1.append(("Spearman ρ (95% CI)",
                        [f_rho(M_TOT["uncal_ix"]), f_rho(M_TOT["ols_ix"]),
                         f_rho(M_TOT["pb_ix"])], False))
        rows_r1.append(("sMAPE (%)",
                        [f_smape(M_TOT["uncal_ix"]), f_smape(M_TOT["ols_ix"]),
                         f_smape(M_TOT["pb_ix"])], False))
        rows_r1.append(("Mean Bias (%)  (Predicted - CMR)",
                        [f_bias_pct(M_TOT["uncal_ix"]), f_bias_pct(M_TOT["ols_ix"]),
                         f_bias_pct(M_TOT["pb_ix"])], False))
        rows_r1.append(("95% LoA (%)",
                        [f_loa_pct(M_TOT["uncal_ix"]), f_loa_pct(M_TOT["ols_ix"]),
                         f_loa_pct(M_TOT["pb_ix"])], False))

    foots_r1 = [SIGN_NOTE, FORMULA_OLS_ABS + " " + FORMULA_PB_ABS]
    if FORMULA_OLS_IX: foots_r1.append(FORMULA_OLS_IX + " " + FORMULA_PB_IX)
    foots_r1.append(ABBR_REL)

    _write_sheet(
        ws, title=f"Table R1. Relative performance in the {cohort_name} cohort (n = {n_total})",
        col_headers=COLS_ABS_TOT, rows=rows_r1, footnotes=foots_r1,
        col_widths=[40, 26, 26, 26],
    )

    if has_lvh and M_POS is not None and M_NEG is not None:
        ws = wb.create_sheet("Rel-LVH-OLS")
        rows_r2 = []
        rows_r2.append(("Uncalibrated - relative (%)", ["", ""], True))
        rows_r2.append(("Spearman ρ (95% CI)",
                        [f_rho(M_POS["uncal_abs"]), f_rho(M_NEG["uncal_abs"])], False))
        rows_r2.append(("sMAPE (%)",
                        [f_smape(M_POS["uncal_abs"]), f_smape(M_NEG["uncal_abs"])], False))
        rows_r2.append(("Mean Bias (%)  (Predicted - CMR)*",
                        [f_bias_pct(M_POS["uncal_abs"]), f_bias_pct(M_NEG["uncal_abs"])], False))
        rows_r2.append(("95% LoA (%)",
                        [f_loa_pct(M_POS["uncal_abs"]), f_loa_pct(M_NEG["uncal_abs"])], False))

        rows_r2.append(("Linear recalibrated - relative (%)", ["", ""], True))
        rows_r2.append(("Spearman ρ (95% CI)",
                        [f_rho(M_POS["ols_abs"]), f_rho(M_NEG["ols_abs"])], False))
        rows_r2.append(("sMAPE (%)",
                        [f_smape(M_POS["ols_abs"]), f_smape(M_NEG["ols_abs"])], False))
        rows_r2.append(("Mean Bias (%)  (Predicted - CMR)†",
                        [f_bias_pct(M_POS["ols_abs"]), f_bias_pct(M_NEG["ols_abs"])], False))
        rows_r2.append(("95% LoA (%)",
                        [f_loa_pct(M_POS["ols_abs"]), f_loa_pct(M_NEG["ols_abs"])], False))

        if has_indexed:
            rows_r2.append(("Linear recalibrated indexed - relative (%)", ["", ""], True))
            rows_r2.append(("Spearman ρ (95% CI)",
                            [f_rho(M_POS["ols_ix"]), f_rho(M_NEG["ols_ix"])], False))
            rows_r2.append(("sMAPE (%)",
                            [f_smape(M_POS["ols_ix"]), f_smape(M_NEG["ols_ix"])], False))
            rows_r2.append(("Mean Bias (%)  (Predicted - CMR)‡",
                            [f_bias_pct(M_POS["ols_ix"]), f_bias_pct(M_NEG["ols_ix"])], False))
            rows_r2.append(("95% LoA (%)",
                            [f_loa_pct(M_POS["ols_ix"]), f_loa_pct(M_NEG["ols_ix"])], False))

        foots_r2 = [
            SIGN_NOTE,
            f"* Bias LVH+ vs LVH- (uncalibrated, absolute reference): Mann-Whitney U, {f_p(mw['uncal_abs'])}.",
            f"† Bias LVH+ vs LVH- (OLS recalibrated, absolute reference): Mann-Whitney U, {f_p(mw['ols_abs'])}.",
        ]
        if has_indexed:
            foots_r2.append(
                f"‡ Bias LVH+ vs LVH- (OLS recalibrated indexed): Mann-Whitney U, {f_p(mw['ols_ix'])}.")
        foots_r2.append(ABBR_REL)

        _write_sheet(
            ws,
            title="Table R2. Relative performance stratified by exercise-induced LVH (linear OLS recalibration)",
            col_headers=COLS_LVH, rows=rows_r2, footnotes=foots_r2,
            col_widths=[42, 28, 28],
        )

        ws = wb.create_sheet("Rel-LVH-PB")
        rows_r3 = []
        rows_r3.append(("Passing-Bablok recalibrated - relative (%)", ["", ""], True))
        rows_r3.append(("Spearman ρ (95% CI)",
                        [f_rho(M_POS["pb_abs"]), f_rho(M_NEG["pb_abs"])], False))
        rows_r3.append(("sMAPE (%)",
                        [f_smape(M_POS["pb_abs"]), f_smape(M_NEG["pb_abs"])], False))
        rows_r3.append(("Mean Bias (%)  (Predicted - CMR)*",
                        [f_bias_pct(M_POS["pb_abs"]), f_bias_pct(M_NEG["pb_abs"])], False))
        rows_r3.append(("95% LoA (%)",
                        [f_loa_pct(M_POS["pb_abs"]), f_loa_pct(M_NEG["pb_abs"])], False))

        if has_indexed:
            rows_r3.append(("Passing-Bablok recalibrated indexed - relative (%)", ["", ""], True))
            rows_r3.append(("Spearman ρ (95% CI)",
                            [f_rho(M_POS["pb_ix"]), f_rho(M_NEG["pb_ix"])], False))
            rows_r3.append(("sMAPE (%)",
                            [f_smape(M_POS["pb_ix"]), f_smape(M_NEG["pb_ix"])], False))
            rows_r3.append(("Mean Bias (%)  (Predicted - CMR)†",
                            [f_bias_pct(M_POS["pb_ix"]), f_bias_pct(M_NEG["pb_ix"])], False))
            rows_r3.append(("95% LoA (%)",
                            [f_loa_pct(M_POS["pb_ix"]), f_loa_pct(M_NEG["pb_ix"])], False))

        foots_r3 = [
            SIGN_NOTE,
            f"* Bias LVH+ vs LVH- (PB recalibrated, absolute reference): Mann-Whitney U, {f_p(mw['pb_abs'])}.",
        ]
        if has_indexed:
            foots_r3.append(
                f"† Bias LVH+ vs LVH- (PB recalibrated indexed): Mann-Whitney U, {f_p(mw['pb_ix'])}.")
        foots_r3.append("Sensitivity analysis: Passing-Bablok is non-parametric and does "
                        "not include sex as a covariate.")
        foots_r3.append(ABBR_REL)

        _write_sheet(
            ws,
            title="Table R3. Relative performance stratified by exercise-induced LVH (Passing-Bablok, sensitivity)",
            col_headers=COLS_LVH, rows=rows_r3, footnotes=foots_r3,
            col_widths=[42, 28, 28],
        )

    out_xlsx = os.path.join(output_dir, "LVM_AI_results_tables.xlsx")
    wb.save(out_xlsx)
    print(f"  Workbook saved: {out_xlsx}")
    for name in wb.sheetnames:
        print(f"    - Sheet '{name}'")

    # --- Table 1 summary sheet ---------------------------------------
    if has_lvh and has_indexed and M_POS is not None and M_NEG is not None:
        ws_main = wb.create_sheet("Table1-Summary", 0)
        wb.active = wb.sheetnames.index("Table1-Summary")

        cols_main = [
            f"Exercise-Induced LVH\n(n = {n_pos})",
            f"No Exercise-Induced LVH\n(n = {n_neg})",
            "P-value*"
        ]

        rows_main = []
        rows_main.append(("Linear Recalibrated (OLS)", ["", "", ""], True))
        rows_main.append(("Spearman ρ (95% CI)",
                          [f_rho(M_POS.get("ols_ix")), f_rho(M_NEG.get("ols_ix")), ""], False))
        rows_main.append(("sMAPE (%)",
                          [f_smape(M_POS.get("ols_ix")), f_smape(M_NEG.get("ols_ix")), ""], False))
        rows_main.append(("Mean Bias (%)  (Predicted - CMR)",
                          [f_bias_pct(M_POS.get("ols_ix")), f_bias_pct(M_NEG.get("ols_ix")),
                           f_p(mw.get('ols_ix', np.nan))], False))
        rows_main.append(("95% LoA (%)",
                          [f_loa_pct(M_POS.get("ols_ix")), f_loa_pct(M_NEG.get("ols_ix")), ""], False))

        rows_main.append(("Passing-Bablok Recalibrated", ["", "", ""], True))
        rows_main.append(("Spearman ρ (95% CI)",
                          [f_rho(M_POS.get("pb_ix")), f_rho(M_NEG.get("pb_ix")), ""], False))
        rows_main.append(("sMAPE (%)",
                          [f_smape(M_POS.get("pb_ix")), f_smape(M_NEG.get("pb_ix")), ""], False))
        rows_main.append(("Mean Bias (%)  (Predicted - CMR)",
                          [f_bias_pct(M_POS.get("pb_ix")), f_bias_pct(M_NEG.get("pb_ix")),
                           f_p(mw.get('pb_ix', np.nan))], False))
        rows_main.append(("95% LoA (%)",
                          [f_loa_pct(M_POS.get("pb_ix")), f_loa_pct(M_NEG.get("pb_ix")), ""], False))

        foots_main = [
            "* P-value calculated using Mann-Whitney U test comparing the relative bias (Predicted - CMR) between LVH+ and LVH- groups.",
            "Note: Positive values indicate AI overestimation; negative values indicate AI underestimation."
        ]

        _write_sheet(
            ws_main,
            title="Table 1. Relative performance of indexed LVM estimates stratified by exercise-induced LVH",
            col_headers=cols_main, rows=rows_main, footnotes=foots_main,
            col_widths=[38, 30, 30, 18],
        )

    wb.save(out_xlsx)

    # --- 10. ROC ANALYSIS ---------------------------------------------
    print(f"\n{'='*64}")
    print(f"  10. ROC ANALYSIS - Regression head as classifier")
    print(f"{'='*64}")

    if not has_lvh:
        print("  No LVH labels available - ROC analysis skipped")
    else:
        sub_roc = results.dropna(subset=["lvh_label_int"]).copy()
        sub_roc["lvh_label_int"] = sub_roc["lvh_label_int"].astype(int)

        if sub_roc["lvh_label_int"].nunique() < 2 or len(sub_roc) < 10:
            print(f"  [WAARSCHUWING] Cohort '{cohort_name}': insufficient data or "
                  f"only one LVH class (n={len(sub_roc)}, "
                  f"classes={sub_roc['lvh_label_int'].nunique()}) - ROC skipped")
        else:
            y = sub_roc["lvh_label_int"].values

            roc_models = []
            roc_models.append(("Raw Absolute LVM",
                               sub_roc["lvm_predicted"].values, "tab:gray"))
            roc_models.append(("OLS Calibrated Absolute LVM",
                               sub_roc["lvm_predicted_calibrated"].values, "tab:blue"))
            roc_models.append(("Passing-Bablok Calibrated Absolute LVM",
                               sub_roc["lvm_predicted_pb"].values, "tab:cyan"))
            if "lvm_predicted_indexed" in sub_roc.columns:
                roc_models.append(("Raw Indexed LVM",
                                   sub_roc["lvm_predicted_indexed"].values, "tab:olive"))
            if "lvm_predicted_indexed_calibrated" in sub_roc.columns:
                roc_models.append(("OLS Calibrated Indexed LVM",
                                   sub_roc["lvm_predicted_indexed_calibrated"].values,
                                   "tab:red"))
            if "lvm_predicted_indexed_pb" in sub_roc.columns:
                roc_models.append(("Passing-Bablok Calibrated Indexed LVM",
                                   sub_roc["lvm_predicted_indexed_pb"].values,
                                   "tab:orange"))

            roc_results = []
            for name, score, color in roc_models:
                mask = np.isfinite(score)
                if mask.sum() < 10 or len(np.unique(y[mask])) < 2:
                    roc_results.append((name, score, color, np.nan, np.nan, np.nan,
                                        None, None))
                    continue
                fpr, tpr, _ = roc_curve(y[mask], score[mask])
                auc_val, auc_lo, auc_hi = bootstrap_auc(y[mask], score[mask])
                roc_results.append((name, score, color, auc_val, auc_lo, auc_hi,
                                    fpr, tpr))
                print(f"  {name:<45}  AUC = {auc_val:.3f}  "
                      f"(95% CI {auc_lo:.3f}-{auc_hi:.3f})")

            fig_roc, ax_roc = plt.subplots(figsize=(9, 9))
            for name, score, color, auc_val, auc_lo, auc_hi, fpr, tpr in roc_results:
                if fpr is None:
                    continue
                label = (f"{name}\n"
                         f"AUC = {auc_val:.2f} (95% CI {auc_lo:.2f}-{auc_hi:.2f})")
                ax_roc.plot(fpr, tpr, color=color, lw=2, label=label)
            ax_roc.plot([0, 1], [0, 1], "--", color="gray", lw=1, label="Chance")
            ax_roc.set_xlabel("1 - Specificity (False Positive Rate)", fontsize=11)
            ax_roc.set_ylabel("Sensitivity (True Positive Rate)", fontsize=11)
            ax_roc.set_title(
                f"[{cohort_name}] ROC Curves: LVM-AI Regression Output as Classifier for "
                "CMR-Defined Exercise-Induced LVH", fontsize=12)
            ax_roc.set_xlim(0, 1)
            ax_roc.set_ylim(0, 1.02)
            ax_roc.legend(loc="lower right", fontsize=9, framealpha=0.95)
            ax_roc.grid(alpha=0.3)
            fig_roc.tight_layout()
            fig_roc.savefig(os.path.join(output_dir, "fig_roc_regression.png"),
                            dpi=150, bbox_inches="tight")
            print(f"\n  ROC figure saved: {output_dir}/fig_roc_regression.png")

            ws_roc = wb.create_sheet("ROC-AUC")
            cols_roc = [f"AUC (95% CI)\nLVH+ (n = {(y==1).sum()}) vs "
                        f"LVH− (n = {(y==0).sum()})"]
            rows_roc = []
            rows_roc.append(("Absolute LVM estimates", [""], True))
            for name, score, color, auc_val, auc_lo, auc_hi, fpr, tpr in roc_results:
                if "Absolute" not in name:
                    continue
                if np.isfinite(auc_val):
                    val = f"{auc_val:.2f} (95% CI {auc_lo:.2f}-{auc_hi:.2f})"
                else:
                    val = "-"
                rows_roc.append((name, [val], False))

            rows_roc.append(("Indexed LVM estimates", [""], True))
            for name, score, color, auc_val, auc_lo, auc_hi, fpr, tpr in roc_results:
                if "Indexed" not in name:
                    continue
                if np.isfinite(auc_val):
                    val = f"{auc_val:.2f} (95% CI {auc_lo:.2f}-{auc_hi:.2f})"
                else:
                    val = "-"
                rows_roc.append((name, [val], False))

            foots_roc = [
                "AUC = area under the receiver operating characteristic curve. "
                "95% confidence intervals derived from 1000 bootstrap iterations.",
                "The regression output (continuous predicted LVM) was used as a "
                "ranking score to discriminate between athletes with and without "
                "CMR-defined exercise-induced LVH. AUC = 0.5 indicates no "
                "discrimination; AUC = 1.0 indicates perfect discrimination.",
                "Abbreviations: AUC, area under the curve; CI, confidence interval; "
                "CMR, cardiovascular magnetic resonance; LVH, left ventricular "
                "hypertrophy; LVM, left ventricular mass; OLS, ordinary least squares."
            ]

            _write_sheet(
                ws_roc,
                title="Table. Discriminative performance of LVM-AI regression "
                      "output for exercise-induced LVH",
                col_headers=cols_roc, rows=rows_roc, footnotes=foots_roc,
                col_widths=[48, 36],
            )

            wb.save(out_xlsx)
            print(f"  ROC-AUC sheet added to workbook")

    # --- Sluit alle figuren van deze cohort om geheugen vrij te maken --
    plt.close('all')


# =========================================================================
# === DRIE COHORTEN UITVOEREN =============================================
# =========================================================================
cohorts = [
    ("Totaal",  full_results.copy(),                                                    os.path.join(BASE_OUTPUT_DIR, "Totaal")),
    ("Mannen",  full_results[full_results["sex"] == 1].reset_index(drop=True).copy(),    os.path.join(BASE_OUTPUT_DIR, "Mannen")),
    ("Vrouwen", full_results[full_results["sex"] == 0].reset_index(drop=True).copy(),    os.path.join(BASE_OUTPUT_DIR, "Vrouwen")),
]

for name, df_cohort, path in cohorts:
    run_analysis(df_cohort, name, path)

print("\n" + "="*72)
print("KLAAR. Alle drie cohort-analyses zijn voltooid.")
for name, df_cohort, path in cohorts:
    print(f"  - {name:<8s} (n={len(df_cohort):3d})  ->  {path}/")
print("="*72)
