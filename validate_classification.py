"""
LVM-AI Classificatie Analyse – LVH detectie bij atleten (SEX-STRATIFIED)
========================================================================
Gebruikt de classificatie-head (output (1,2), index[1] = P(LVH)).
Vergelijkt P(LVH) tegen ground-truth lvh_label.

Refactored: de volledige analyse-pipeline wordt VIER keer uitgevoerd:
  1. Totale cohort        -> classificationresult/Totaal/
  2. Mannen (sex=1)       -> classificationresult/Mannen/
  3. Vrouwen (sex=0)      -> classificationresult/Vrouwen/
  4. Genormaliseerd       -> classificationresult/Genormaliseerd/
     (pooled cohort; LVM-AI gebruikt sex als model-input, dus de output is
      intrinsiek sex-aware. ECG-criteria behouden hun klinische
      sex-specifieke drempels.)

NIEUW T.O.V. DE OORSPRONKELIJKE STRATIFIED VERSIE:
  - run_analysis(results, cohort_name, output_dir, is_normalized=False) als
    master-functie die alles na de inferentie omvat.
  - Per cohort eigen subdirectory, eigen ROC-figuur, eigen CSV en eigen
    Excel workbook.
  - Binnen elke cohort blijft de Male-vs-Female side-by-side evaluatie
    intact; voor de Mannen/Vrouwen cohorten wordt de afwezige sex-cel
    gracefully leeg gelaten ("Onvoldoende data").

Gebruik: python validate_classification.py
"""

import subprocess, sys

def install(packages):
    for pkg in packages:
        try:
            __import__(pkg.split("==")[0].replace("-","_").replace("beautifulsoup4","bs4"))
        except ImportError:
            print(f"Installeren: {pkg}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

install([
    "tensorflow==2.19.0",
    "beautifulsoup4", "lxml",
    "numpy", "pandas", "scikit-learn",
    "matplotlib", "seaborn", "scipy",
    "openpyxl",
])

import os
import base64
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from scipy.signal import find_peaks
from sklearn.metrics import (
    roc_auc_score, roc_curve, confusion_matrix,
    accuracy_score, brier_score_loss,
)
from sklearn.calibration import calibration_curve

warnings.filterwarnings("ignore")

# ─── CONFIGURATIE ─────────────────────────────────────────────────────────
MODEL_FILE = "ecg_rest_raw_age_sex_bmi_lvm_asymmetric_loss.h5"
ECG_DIR    = "ecg_bestanden"
LABELS_CSV = "regressionlabels.csv"
BASE_OUTPUT_DIR = "classificationresult"
os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)

# Normalisatie-constanten uit ml4h repo
ECG_NORM = 2000.0
AGE_MEAN, AGE_STD = 63.35798891483556, 7.554638350423902
BMI_MEAN, BMI_STD = 27.3397, 4.7721
LEAD_ORDER  = ["I","II","III","V1","V2","V3","V4","V5","V6","aVF","aVL","aVR"]
ECG_SAMPLES = 5000

TARGET_SPEC = 0.95   # doel-specificiteit voor globaal AI operating point

# ─── SOKOLOW-LYON CONFIGURATIE ───────────────────────────────────────────
SL_THRESHOLD_UV   = 3500.0
ECG_SAMPLE_RATE   = 500
SL_LEADS_R        = ["V5", "V6"]
SL_LEAD_S         = "V1"

# ─── SEX-MAPPING + CRITERIA-DREMPELS ─────────────────────────────────────
MALE_SEX_VALUE   = 1
FEMALE_SEX_VALUE = 0

# Cornell Voltage: R(aVL) + S(V3),  strict ">"
CORNELL_THRESH_MALE_UV   = 2800.0
CORNELL_THRESH_FEMALE_UV = 2000.0

# Peguero-Lo Presti: S(V4) + SD,    "≥"
PEGUERO_THRESH_MALE_UV   = 2800.0
PEGUERO_THRESH_FEMALE_UV = 2300.0

# ─── MODEL LADEN ─────────────────────────────────────────────────────────
print("Model laden...")
import tensorflow as tf
model = tf.keras.models.load_model(MODEL_FILE, compile=False)
print(f"  Model geladen: {MODEL_FILE}")
print(f"  Outputs: {[o.shape for o in model.outputs]}")

# ─── LABELS LADEN ────────────────────────────────────────────────────────
labels_df = pd.read_csv(LABELS_CSV, sep=";", dtype={"sample_id": str})
labels_df["sample_id"] = labels_df["sample_id"].str.strip()
labels_df.columns = [c.strip() for c in labels_df.columns]

lvh_col = next((c for c in ["lvh_label","LVH_label","LVH","lvh","lvh_status","EILVH","eilvh"]
                if c in labels_df.columns), None)
if lvh_col is None:
    sys.exit("FOUT: geen LVH-label kolom gevonden in CSV.")

for col in ["age","sex","bmi","lvm_grams"]:
    if col in labels_df.columns:
        labels_df[col] = labels_df[col].astype(str).str.replace(",", ".").astype(float)

def parse_lvh(x):
    if pd.isna(x): return np.nan
    s = str(x).strip().lower()
    if s in {"1","1.0","ja","yes","y","true","t","lvh","lvh+","positive","pos"}: return 1
    if s in {"0","0.0","nee","no","n","false","f","geen","geen lvh","geen_lvh",
             "no lvh","no_lvh","lvh-","negative","neg","control"}: return 0
    try:
        v = float(s.replace(",", "."))
        if v in (0, 1): return int(v)
    except ValueError: pass
    return np.nan

labels_df["lvh_label_int"] = labels_df[lvh_col].apply(parse_lvh)
print(f"\nLabels geladen: {len(labels_df)} atleten")
print(f"  LVH-kolom in CSV     : '{lvh_col}'")
print(f"  Na parse (0/1/NaN)   : {labels_df['lvh_label_int'].value_counts(dropna=False).to_dict()}")

# ─── XML LOADER ──────────────────────────────────────────────────────────
def decode_b64(raw, scale=1.0):
    return np.frombuffer(base64.b64decode(raw), dtype="<i2").astype(np.float32) * scale

def load_xml(path):
    import bs4
    with open(path, "r", errors="replace") as f:
        soup = bs4.BeautifulSoup(f, "lxml")
    v = {}
    for wf in soup.find_all("waveform"):
        wt = wf.find("waveformtype")
        if wt is None or wt.text.strip() != "Rhythm": continue
        for ld in wf.find_all("leaddata"):
            lid = ld.find("leadid").text.strip()
            sc_tag = ld.find("leadamplitudeunitsperbit")
            sc = float(sc_tag.text.replace(",", ".")) if sc_tag else 1.0
            v[lid] = decode_b64(ld.find("waveformdata").text.strip(), sc)
        break
    if "III" not in v: v["III"] = v["II"] - v["I"]
    if "aVR" not in v: v["aVR"] = -(v["I"] + v["II"]) / 2
    if "aVL" not in v: v["aVL"] =  v["I"] - v["II"] / 2
    if "aVF" not in v: v["aVF"] =  v["II"] - v["I"] / 2
    def resample(x):
        n = len(x)
        return x if n == ECG_SAMPLES else np.interp(np.linspace(0, n, ECG_SAMPLES), np.arange(n), x)
    return np.column_stack([resample(v[l]) for l in LEAD_ORDER]).astype(np.float32)

# ─── PEAK DETECTION ──────────────────────────────────────────────────────
def _detrend(signal):
    n = len(signal); x = np.arange(n)
    coeffs = np.polyfit(x, signal, 1)
    return signal - np.polyval(coeffs, x)

def detect_qrs_amplitude(signal_uv, polarity="positive", fs=ECG_SAMPLE_RATE):
    sig = _detrend(np.asarray(signal_uv, dtype=np.float64))
    search_signal = sig if polarity == "positive" else -sig
    abs_max = np.max(np.abs(sig))
    if abs_max <= 0:
        return 0.0
    peaks, _ = find_peaks(search_signal, height=0.3 * abs_max,
                          distance=int(0.4 * fs))
    if len(peaks) == 0:
        return float(abs_max)
    return float(np.median(np.abs(sig[peaks])))

def sokolow_lyon(ecg_array, lead_order=LEAD_ORDER, fs=ECG_SAMPLE_RATE):
    iV1 = lead_order.index("V1"); iV5 = lead_order.index("V5"); iV6 = lead_order.index("V6")
    s_v1 = detect_qrs_amplitude(ecg_array[:, iV1], polarity="negative", fs=fs)
    r_v5 = detect_qrs_amplitude(ecg_array[:, iV5], polarity="positive", fs=fs)
    r_v6 = detect_qrs_amplitude(ecg_array[:, iV6], polarity="positive", fs=fs)
    return s_v1 + max(r_v5, r_v6), dict(s_v1=s_v1, r_v5=r_v5, r_v6=r_v6, r_max=max(r_v5, r_v6))

def cornell_voltage(ecg_array, lead_order=LEAD_ORDER, fs=ECG_SAMPLE_RATE):
    iaVL = lead_order.index("aVL"); iV3 = lead_order.index("V3")
    r_avl = detect_qrs_amplitude(ecg_array[:, iaVL], polarity="positive", fs=fs)
    s_v3  = detect_qrs_amplitude(ecg_array[:, iV3],  polarity="negative", fs=fs)
    return r_avl + s_v3, dict(r_avl=r_avl, s_v3=s_v3)

def peguero_lo_presti(ecg_array, lead_order=LEAD_ORDER, fs=ECG_SAMPLE_RATE):
    iV4 = lead_order.index("V4")
    s_v4 = detect_qrs_amplitude(ecg_array[:, iV4], polarity="negative", fs=fs)
    s_per_lead = {lead: detect_qrs_amplitude(ecg_array[:, k], polarity="negative", fs=fs)
                  for k, lead in enumerate(lead_order)}
    sd_lead = max(s_per_lead, key=s_per_lead.get)
    sd_uv   = s_per_lead[sd_lead]
    return s_v4 + sd_uv, dict(s_v4=s_v4, sd_uv=sd_uv, sd_lead=sd_lead)


# ─── INFERENTIE ──────────────────────────────────────────────────────────
print("\nInferentie starten...")
ecg_index = {Path(p).stem: str(p) for p in Path(ECG_DIR).glob("*.xml")}
print(f"  ECG-bestanden gevonden: {len(ecg_index)}")

n_inputs = len(model.inputs)
records, fouten = [], 0
for i, row in labels_df.iterrows():
    sid = str(row["sample_id"])
    if sid not in ecg_index: fouten += 1; continue
    if pd.isna(row["lvh_label_int"]):
        fouten += 1; continue
    try:
        ecg  = load_xml(ecg_index[sid])
        sex_val = int(row["sex"])

        # Sokolow-Lyon
        sl_uv, sl_c = sokolow_lyon(ecg)
        sl_pred = int(sl_uv >= SL_THRESHOLD_UV)

        # Cornell (sex-specific)
        cornell_uv, cornell_c = cornell_voltage(ecg)
        cornell_thr = CORNELL_THRESH_MALE_UV if sex_val == MALE_SEX_VALUE else (
                      CORNELL_THRESH_FEMALE_UV if sex_val == FEMALE_SEX_VALUE else CORNELL_THRESH_MALE_UV)
        cornell_pred = int(cornell_uv > cornell_thr)

        # Peguero-Lo Presti (sex-specific)
        peguero_uv, peg_c = peguero_lo_presti(ecg)
        peguero_thr = PEGUERO_THRESH_MALE_UV if sex_val == MALE_SEX_VALUE else (
                      PEGUERO_THRESH_FEMALE_UV if sex_val == FEMALE_SEX_VALUE else PEGUERO_THRESH_MALE_UV)
        peguero_pred = int(peguero_uv >= peguero_thr)

        # AI
        norm = (ecg / ECG_NORM)[np.newaxis, ...]
        age_n = np.array([[(float(row["age"]) - AGE_MEAN) / AGE_STD]], dtype=np.float32)
        bmi_n = np.array([[(float(row["bmi"]) - BMI_MEAN) / BMI_STD]], dtype=np.float32)
        sex = np.array([[1 - sex_val, sex_val]], dtype=np.float32)

        if n_inputs == 4:   out = model.predict([norm, age_n, sex, bmi_n], verbose=0)
        elif n_inputs == 2: out = model.predict([norm, np.array([[age_n[0,0], sex_val, bmi_n[0,0]]], dtype=np.float32)], verbose=0)
        else:               out = model.predict(norm, verbose=0)

        p_lvh = None
        if isinstance(out, (list, tuple)):
            for o in out:
                if o.shape[-1] == 2:
                    p_lvh = float(o[0, 1]); break
        if p_lvh is None: fouten += 1; continue

        records.append({
            "sample_id":         sid,
            "age":               float(row["age"]),
            "sex":               sex_val,
            "bmi":               float(row["bmi"]),
            "lvh_label":         int(row["lvh_label_int"]),
            "p_lvh":             p_lvh,
            "p_no_lvh":          1.0 - p_lvh,
            "sokolow_lyon_uv":   sl_uv,
            "sokolow_lyon_mv":   sl_uv / 1000.0,
            "sokolow_lyon_pred": sl_pred,
            "sl_s_v1_uv":        sl_c["s_v1"],
            "sl_r_v5_uv":        sl_c["r_v5"],
            "sl_r_v6_uv":        sl_c["r_v6"],
            "cornell_uv":        cornell_uv,
            "cornell_mv":        cornell_uv / 1000.0,
            "cornell_thr_uv":    cornell_thr,
            "cornell_pred":      cornell_pred,
            "cornell_r_avl_uv":  cornell_c["r_avl"],
            "cornell_s_v3_uv":   cornell_c["s_v3"],
            "peguero_uv":        peguero_uv,
            "peguero_mv":        peguero_uv / 1000.0,
            "peguero_thr_uv":    peguero_thr,
            "peguero_pred":      peguero_pred,
            "peguero_s_v4_uv":   peg_c["s_v4"],
            "peguero_sd_uv":     peg_c["sd_uv"],
            "peguero_sd_lead":   peg_c["sd_lead"],
        })
        if (i + 1) % 50 == 0: print(f"  {i+1}/{len(labels_df)} verwerkt...")
    except Exception as e:
        print(f"  FOUT bij {sid}: {e}"); fouten += 1

full_results = pd.DataFrame(records)
print(f"\nKlaar: {len(full_results)} voorspellingen, {fouten} fouten/overgeslagen")
if len(full_results) < 6 or full_results["lvh_label"].nunique() < 2:
    sys.exit("Te weinig data of slechts één klasse — analyse niet mogelijk.")


# ═════════════════════════════════════════════════════════════════════════
# ─── HELPERS VOOR STATISTIEK ─────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════
def classification_stats_from_pred(y_true, y_pred):
    y_true = np.asarray(y_true).astype(int)
    y_pred = np.asarray(y_pred).astype(int)
    if len(y_true) == 0:
        return dict(tn=0, fp=0, fn=0, tp=0,
                    sens=np.nan, spec=np.nan, ppv=np.nan, npv=np.nan, accuracy=np.nan)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    sens = tp / (tp + fn) if (tp + fn) else np.nan
    spec = tn / (tn + fp) if (tn + fp) else np.nan
    ppv  = tp / (tp + fp) if (tp + fp) else np.nan
    npv  = tn / (tn + fn) if (tn + fn) else np.nan
    acc  = (tp + tn) / (tp + tn + fp + fn) if (tp + tn + fp + fn) else np.nan
    return dict(tn=int(tn), fp=int(fp), fn=int(fn), tp=int(tp),
                sens=sens, spec=spec, ppv=ppv, npv=npv, accuracy=acc)

def bootstrap_auc_ci(y_true, y_score, n_boot=1000, seed=42):
    y_true = np.asarray(y_true).astype(int)
    y_score = np.asarray(y_score).astype(float)
    if len(y_true) == 0 or len(np.unique(y_true)) < 2:
        return np.nan, np.nan
    rng = np.random.default_rng(seed)
    boot = []
    for _ in range(n_boot):
        idx = rng.integers(0, len(y_true), len(y_true))
        if len(np.unique(y_true[idx])) < 2: continue
        boot.append(roc_auc_score(y_true[idx], y_score[idx]))
    if not boot: return np.nan, np.nan
    return float(np.percentile(boot, 2.5)), float(np.percentile(boot, 97.5))

def safe_auc(y_true, y_score):
    y_true = np.asarray(y_true).astype(int)
    if len(y_true) == 0 or len(np.unique(y_true)) < 2:
        return np.nan
    return float(roc_auc_score(y_true, y_score))


# ═════════════════════════════════════════════════════════════════════════
# ─── STRATUM EVALUATIE ───────────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════
def evaluate_stratum(df, label, youden_thr, spec95_thr):
    """Alle metrics voor één sex-stratum. Globale drempels worden meegegeven."""
    empty = dict(tp=0, tn=0, fp=0, fn=0,
                 sens=np.nan, spec=np.nan, ppv=np.nan, npv=np.nan, accuracy=np.nan)
    if len(df) == 0 or df["lvh_label"].nunique() < 2:
        print(f"  WAARSCHUWING: stratum '{label}' heeft te weinig data of slechts één klasse.")
        return dict(
            n=len(df), n_pos=int((df["lvh_label"]==1).sum()) if len(df) else 0,
            n_neg=int((df["lvh_label"]==0).sum()) if len(df) else 0,
            ai_auc=np.nan, ai_auc_lo=np.nan, ai_auc_hi=np.nan,
            ai_youden=empty.copy(), ai_spec95=empty.copy(),
            sl_auc=np.nan, sl_auc_lo=np.nan, sl_auc_hi=np.nan, sl=empty.copy(),
            cornell_auc=np.nan, cornell_auc_lo=np.nan, cornell_auc_hi=np.nan, cornell=empty.copy(),
            peguero_auc=np.nan, peguero_auc_lo=np.nan, peguero_auc_hi=np.nan, peguero=empty.copy(),
            fpr_ai=None, tpr_ai=None, fpr_sl=None, tpr_sl=None,
            fpr_cornell=None, tpr_cornell=None, fpr_peguero=None, tpr_peguero=None,
        )

    yt = df["lvh_label"].values.astype(int)

    # ── LVM-AI ──
    ys_ai = df["p_lvh"].values.astype(float)
    ai_auc = safe_auc(yt, ys_ai)
    ai_auc_lo, ai_auc_hi = bootstrap_auc_ci(yt, ys_ai)
    fpr_ai, tpr_ai, _ = roc_curve(yt, ys_ai)

    pred_youden = (ys_ai >= youden_thr).astype(int)
    ai_youden = classification_stats_from_pred(yt, pred_youden)
    if not np.isnan(spec95_thr):
        pred_spec95 = (ys_ai >= spec95_thr).astype(int)
        ai_spec95 = classification_stats_from_pred(yt, pred_spec95)
    else:
        ai_spec95 = empty.copy()

    # ── Sokolow-Lyon ──
    ys_sl = df["sokolow_lyon_uv"].values.astype(float)
    sl_auc = safe_auc(yt, ys_sl)
    sl_auc_lo, sl_auc_hi = bootstrap_auc_ci(yt, ys_sl)
    fpr_sl, tpr_sl, _ = roc_curve(yt, ys_sl)
    sl = classification_stats_from_pred(yt, df["sokolow_lyon_pred"].values.astype(int))

    # ── Cornell ──
    ys_c = df["cornell_uv"].values.astype(float)
    cornell_auc = safe_auc(yt, ys_c)
    cornell_auc_lo, cornell_auc_hi = bootstrap_auc_ci(yt, ys_c)
    fpr_cornell, tpr_cornell, _ = roc_curve(yt, ys_c)
    cornell = classification_stats_from_pred(yt, df["cornell_pred"].values.astype(int))

    # ── Peguero-Lo Presti ──
    ys_p = df["peguero_uv"].values.astype(float)
    peguero_auc = safe_auc(yt, ys_p)
    peguero_auc_lo, peguero_auc_hi = bootstrap_auc_ci(yt, ys_p)
    fpr_peguero, tpr_peguero, _ = roc_curve(yt, ys_p)
    peguero = classification_stats_from_pred(yt, df["peguero_pred"].values.astype(int))

    return dict(
        n=len(df), n_pos=int((yt==1).sum()), n_neg=int((yt==0).sum()),
        ai_auc=ai_auc, ai_auc_lo=ai_auc_lo, ai_auc_hi=ai_auc_hi,
        ai_youden=ai_youden, ai_spec95=ai_spec95,
        sl_auc=sl_auc, sl_auc_lo=sl_auc_lo, sl_auc_hi=sl_auc_hi, sl=sl,
        cornell_auc=cornell_auc, cornell_auc_lo=cornell_auc_lo, cornell_auc_hi=cornell_auc_hi, cornell=cornell,
        peguero_auc=peguero_auc, peguero_auc_lo=peguero_auc_lo, peguero_auc_hi=peguero_auc_hi, peguero=peguero,
        fpr_ai=fpr_ai, tpr_ai=tpr_ai,
        fpr_sl=fpr_sl, tpr_sl=tpr_sl,
        fpr_cornell=fpr_cornell, tpr_cornell=tpr_cornell,
        fpr_peguero=fpr_peguero, tpr_peguero=tpr_peguero,
    )


# ═════════════════════════════════════════════════════════════════════════
# ─── PRINT/PLOT/EXCEL HELPERS ────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════
def fmt_ci(auc_val, lo, hi):
    if np.isnan(auc_val): return "n/a"
    if np.isnan(lo) or np.isnan(hi): return f"{auc_val:.3f}"
    return f"{auc_val:.3f} [{lo:.3f}, {hi:.3f}]"

def fmt_auc(auc_val, lo, hi):
    if np.isnan(auc_val): return "n/a"
    if np.isnan(lo) or np.isnan(hi): return f"{auc_val:.2f}"
    return f"{auc_val:.2f} ({lo:.2f}–{hi:.2f})"

def fmt_pct(x):
    if x is None or (isinstance(x, float) and np.isnan(x)): return "n/a"
    return f"{100*x:.1f}"

def fmt_cm(s):
    return f"{int(s['tp'])} / {int(s['tn'])} / {int(s['fp'])} / {int(s['fn'])}"


def print_stratum(label, r, youden_thr, spec95_thr):
    print(f"\n{'='*64}")
    print(f"  STRATUM: {label}  (n={r['n']}, LVH+ n={r['n_pos']}, LVH- n={r['n_neg']})")
    print(f"{'='*64}")
    if r['n'] == 0 or r['n_pos'] == 0 or r['n_neg'] == 0:
        print("  Onvoldoende data voor dit stratum.")
        return
    print(f"  LVM-AI AUC                : {fmt_ci(r['ai_auc'], r['ai_auc_lo'], r['ai_auc_hi'])}")
    s = r['ai_youden']
    print(f"  LVM-AI @ Youden ({youden_thr:.3f})")
    print(f"    TP={s['tp']}  TN={s['tn']}  FP={s['fp']}  FN={s['fn']}  "
          f"Sens={s['sens']:.3f}  Spec={s['spec']:.3f}")
    s = r['ai_spec95']
    spec_thr_str = f"{spec95_thr:.3f}" if not np.isnan(spec95_thr) else "n/a"
    print(f"  LVM-AI @ 95% Spec ({spec_thr_str})")
    if not np.isnan(s['sens']):
        print(f"    TP={s['tp']}  TN={s['tn']}  FP={s['fp']}  FN={s['fn']}  "
              f"Sens={s['sens']:.3f}  Spec={s['spec']:.3f}")
    else:
        print(f"    (geen Spec95-drempel beschikbaar)")

    print(f"\n  Sokolow-Lyon AUC          : {fmt_ci(r['sl_auc'], r['sl_auc_lo'], r['sl_auc_hi'])}")
    s = r['sl']
    print(f"    TP={s['tp']}  TN={s['tn']}  FP={s['fp']}  FN={s['fn']}  "
          f"Sens={s['sens']:.3f}  Spec={s['spec']:.3f}")

    print(f"\n  Cornell AUC               : {fmt_ci(r['cornell_auc'], r['cornell_auc_lo'], r['cornell_auc_hi'])}")
    s = r['cornell']
    print(f"    TP={s['tp']}  TN={s['tn']}  FP={s['fp']}  FN={s['fn']}  "
          f"Sens={s['sens']:.3f}  Spec={s['spec']:.3f}")

    print(f"\n  Peguero-Lo Presti AUC     : {fmt_ci(r['peguero_auc'], r['peguero_auc_lo'], r['peguero_auc_hi'])}")
    s = r['peguero']
    print(f"    TP={s['tp']}  TN={s['tn']}  FP={s['fp']}  FN={s['fn']}  "
          f"Sens={s['sens']:.3f}  Spec={s['spec']:.3f}")


COLORS = dict(ai="darkorange", sl="seagreen", cornell="royalblue", peguero="crimson")
MARKERS = dict(youden="o", spec95="s", sl_op="D", cornell_op="P", peguero_op="X")

def plot_panel(ax, r, title):
    if r["n"] == 0 or r["n_pos"] == 0 or r["n_neg"] == 0:
        ax.text(0.5, 0.5, f"Onvoldoende data\nvoor {title}",
                ha="center", va="center", transform=ax.transAxes, fontsize=12)
        ax.set_title(title); return

    ax.plot(r["fpr_ai"], r["tpr_ai"], color=COLORS["ai"], lw=2,
            label=f"LVM-AI: AUC={r['ai_auc']:.3f} [{r['ai_auc_lo']:.3f}, {r['ai_auc_hi']:.3f}]")
    ax.plot(r["fpr_sl"], r["tpr_sl"], color=COLORS["sl"], lw=1.8,
            label=f"Sokolow-Lyon: AUC={r['sl_auc']:.3f} [{r['sl_auc_lo']:.3f}, {r['sl_auc_hi']:.3f}]")
    ax.plot(r["fpr_cornell"], r["tpr_cornell"], color=COLORS["cornell"], lw=1.8,
            label=f"Cornell: AUC={r['cornell_auc']:.3f} [{r['cornell_auc_lo']:.3f}, {r['cornell_auc_hi']:.3f}]")
    ax.plot(r["fpr_peguero"], r["tpr_peguero"], color=COLORS["peguero"], lw=1.8,
            label=f"Peguero-LP: AUC={r['peguero_auc']:.3f} [{r['peguero_auc_lo']:.3f}, {r['peguero_auc_hi']:.3f}]")
    ax.plot([0, 1], [0, 1], "--", color="gray", alpha=0.6)

    s = r["ai_youden"]
    if not np.isnan(s["sens"]):
        ax.plot(1 - s["spec"], s["sens"], MARKERS["youden"], color="red", markersize=9,
                label=f"AI Youden: Sens={s['sens']:.2f}, Spec={s['spec']:.2f}")
    s = r["ai_spec95"]
    if not np.isnan(s["sens"]):
        ax.plot(1 - s["spec"], s["sens"], MARKERS["spec95"], color="purple", markersize=9,
                label=f"AI Spec≥95%: Sens={s['sens']:.2f}, Spec={s['spec']:.2f}")
    s = r["sl"]
    if not np.isnan(s["sens"]):
        ax.plot(1 - s["spec"], s["sens"], MARKERS["sl_op"], color="darkgreen", markersize=10,
                label=f"SL @ 3.5 mV: Sens={s['sens']:.2f}, Spec={s['spec']:.2f}")
    s = r["cornell"]
    if not np.isnan(s["sens"]):
        ax.plot(1 - s["spec"], s["sens"], MARKERS["cornell_op"], color="navy", markersize=11,
                label=f"Cornell: Sens={s['sens']:.2f}, Spec={s['spec']:.2f}")
    s = r["peguero"]
    if not np.isnan(s["sens"]):
        ax.plot(1 - s["spec"], s["sens"], MARKERS["peguero_op"], color="darkred", markersize=11,
                label=f"Peguero: Sens={s['sens']:.2f}, Spec={s['spec']:.2f}")

    ax.set_xlabel("1 - Specificity (FPR)")
    ax.set_ylabel("Sensitivity (TPR)")
    ax.set_title(title)
    ax.legend(loc="lower right", fontsize=7)
    ax.set_xlim([-0.02, 1.02]); ax.set_ylim([-0.02, 1.02])
    ax.grid(alpha=0.3)


def label_outcome(y_true, y_pred):
    out = np.empty(len(y_true), dtype=object)
    out[(y_true == 1) & (y_pred == 1)] = "TP"
    out[(y_true == 0) & (y_pred == 1)] = "FP"
    out[(y_true == 0) & (y_pred == 0)] = "TN"
    out[(y_true == 1) & (y_pred == 0)] = "FN"
    return out


# ═════════════════════════════════════════════════════════════════════════
# ─── EXCEL EXPORT ────────────────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def write_excel(xlsx_path, cohort_name, male_r, female_r,
                youden_thr, spec95_thr, NORM_NOTE=""):
    def row_for(name, auc_val, auc_lo, auc_hi, stat):
        return {
            "name":  name,
            "auc":   fmt_auc(auc_val, auc_lo, auc_hi),
            "sens":  fmt_pct(stat["sens"]),
            "spec":  fmt_pct(stat["spec"]),
            "cm":    fmt_cm(stat),
        }

    male_rows = [
        row_for("LVM-AI (Optimal Youden)",      male_r["ai_auc"], male_r["ai_auc_lo"], male_r["ai_auc_hi"], male_r["ai_youden"]),
        row_for("LVM-AI (Predefined 95% Spec)", male_r["ai_auc"], male_r["ai_auc_lo"], male_r["ai_auc_hi"], male_r["ai_spec95"]),
        row_for("Sokolow-Lyon (≥ 3.5 mV)", male_r["sl_auc"], male_r["sl_auc_lo"], male_r["sl_auc_hi"], male_r["sl"]),
        row_for("Cornell Voltage",              male_r["cornell_auc"], male_r["cornell_auc_lo"], male_r["cornell_auc_hi"], male_r["cornell"]),
        row_for("Peguero-Lo Presti",            male_r["peguero_auc"], male_r["peguero_auc_lo"], male_r["peguero_auc_hi"], male_r["peguero"]),
    ]
    female_rows = [
        row_for("LVM-AI (Optimal Youden)",      female_r["ai_auc"], female_r["ai_auc_lo"], female_r["ai_auc_hi"], female_r["ai_youden"]),
        row_for("LVM-AI (Predefined 95% Spec)", female_r["ai_auc"], female_r["ai_auc_lo"], female_r["ai_auc_hi"], female_r["ai_spec95"]),
        row_for("Sokolow-Lyon (≥ 3.5 mV)", female_r["sl_auc"], female_r["sl_auc_lo"], female_r["sl_auc_hi"], female_r["sl"]),
        row_for("Cornell Voltage",              female_r["cornell_auc"], female_r["cornell_auc_lo"], female_r["cornell_auc_hi"], female_r["cornell"]),
        row_for("Peguero-Lo Presti",            female_r["peguero_auc"], female_r["peguero_auc_lo"], female_r["peguero_auc_hi"], female_r["peguero"]),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 2"

    FONT_NAME      = "Arial"
    title_font     = Font(name=FONT_NAME, size=11, bold=True)
    group_font     = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    header_font    = Font(name=FONT_NAME, size=10, bold=True)
    body_font      = Font(name=FONT_NAME, size=10)
    footnote_font  = Font(name=FONT_NAME, size=9)
    abbrev_font    = Font(name=FONT_NAME, size=9, italic=True, bold=True)
    center         = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left           = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin           = Side(border_style="thin",   color="000000")
    thick          = Side(border_style="medium", color="000000")
    divider        = Side(border_style="medium", color="000000")
    shade_fill     = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    male_band      = PatternFill("solid", start_color="2E5C8A", end_color="2E5C8A")
    female_band    = PatternFill("solid", start_color="A23B72", end_color="A23B72")

    TOTAL_COLS = 9

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(
        row=1, column=1,
        value=f"Table 2 [{cohort_name}]: Sex-Stratified Classification Performance "
              "of the LVM-AI and Classical ECG Criteria for Exercise-Induced LVH*")
    title_cell.font = title_font
    title_cell.alignment = left

    group_row = 2
    ws.cell(row=group_row, column=1, value="")
    ws.merge_cells(start_row=group_row, start_column=2, end_row=group_row, end_column=5)
    m_cell = ws.cell(row=group_row, column=2,
                     value=f"Male Athletes (n={male_r['n']}; LVH+ n={male_r['n_pos']}, "
                           f"LVH− n={male_r['n_neg']})")
    m_cell.font = group_font; m_cell.alignment = center; m_cell.fill = male_band

    ws.merge_cells(start_row=group_row, start_column=6, end_row=group_row, end_column=9)
    f_cell = ws.cell(row=group_row, column=6,
                     value=f"Female Athletes (n={female_r['n']}; LVH+ n={female_r['n_pos']}, "
                           f"LVH− n={female_r['n_neg']})")
    f_cell.font = group_font; f_cell.alignment = center; f_cell.fill = female_band

    header_row = 3
    sub_headers = ["Model / Criterion",
                   "AUC (95% CI)", "Sensitivity (%)", "Specificity (%)", "TP / TN / FP / FN",
                   "AUC (95% CI)", "Sensitivity (%)", "Specificity (%)", "TP / TN / FP / FN"]
    for j, txt in enumerate(sub_headers, start=1):
        c = ws.cell(row=header_row, column=j, value=txt)
        c.font = header_font
        c.alignment = center
        c.border = Border(top=thick, bottom=thin)

    body_start = header_row + 1
    for i, (mrow, frow) in enumerate(zip(male_rows, female_rows)):
        r = body_start + i
        shaded = (i % 2 == 1)
        values = [mrow["name"],
                  mrow["auc"], mrow["sens"], mrow["spec"], mrow["cm"],
                  frow["auc"], frow["sens"], frow["spec"], frow["cm"]]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = body_font
            c.alignment = left if j == 1 else center
            if shaded:
                c.fill = shade_fill

    last_body_row = body_start + len(male_rows) - 1
    for j in range(1, TOTAL_COLS + 1):
        ws.cell(row=last_body_row, column=j).border = Border(bottom=thick)

    for r in range(group_row, last_body_row + 1):
        cell_left  = ws.cell(row=r, column=5)
        cell_right = ws.cell(row=r, column=6)
        existing_l = cell_left.border
        existing_r = cell_right.border
        cell_left.border = Border(
            left=existing_l.left, right=divider,
            top=existing_l.top, bottom=existing_l.bottom)
        cell_right.border = Border(
            left=divider, right=existing_r.right,
            top=existing_r.top, bottom=existing_r.bottom)

    spec_thr_str = f"{spec95_thr:.3f}" if not np.isnan(spec95_thr) else "n/a"
    footnote_row = last_body_row + 2
    ws.merge_cells(start_row=footnote_row, start_column=1,
                   end_row=footnote_row, end_column=TOTAL_COLS)
    fc = ws.cell(
        row=footnote_row, column=1,
        value=f"* Cohort: {cohort_name}. Exercise-induced LVH served as the ground-truth label. "
              "Analyses are stratified by sex; AUC, sensitivity, specificity, and confusion-matrix "
              "counts (TP / TN / FP / FN) are computed independently within male and female strata. "
              f"For LVM-AI, the Optimal Youden threshold ({youden_thr:.3f}) and the Predefined 95% "
              f"Specificity threshold ({spec_thr_str}) were derived on the cohort's pooled data and "
              "subsequently applied within each stratum. AUC 95% CIs were estimated with 1000 "
              "bootstrap resamples within each stratum. ECG-criterion amplitudes were computed as "
              "median QRS-peak amplitudes detected with scipy.signal.find_peaks on the raw ECG "
              "signal. Sokolow-Lyon = |S(V1)| + max(R(V5), R(V6)), threshold ≥ 3.5 mV "
              "(applied to both sexes). Cornell Voltage = R(aVL) + S(V3), threshold > 2.8 mV "
              "(men) / > 2.0 mV (women). Peguero-Lo Presti = S(V4) + Sᴅ (deepest S-wave "
              "across 12 leads), threshold ≥ 2.8 mV (men) / ≥ 2.3 mV (women).")
    fc.font = footnote_font
    fc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[footnote_row].height = 140

    extra_offset = 1
    if NORM_NOTE:
        norm_row = footnote_row + extra_offset
        ws.merge_cells(start_row=norm_row, start_column=1,
                       end_row=norm_row, end_column=TOTAL_COLS)
        nc = ws.cell(row=norm_row, column=1, value=NORM_NOTE)
        nc.font = footnote_font
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[norm_row].height = 50
        extra_offset += 1

    abbrev_row = footnote_row + extra_offset
    ws.merge_cells(start_row=abbrev_row, start_column=1,
                   end_row=abbrev_row, end_column=TOTAL_COLS)
    ac = ws.cell(row=abbrev_row, column=1,
                 value="Abbreviations: AI, artificial intelligence; AUC, area under the ROC curve; "
                       "CI, confidence interval; CMR, cardiac magnetic resonance; ECG, electrocardiogram; "
                       "FN, false negatives; FP, false positives; LVH, left ventricular hypertrophy; "
                       "LVM, left ventricular mass; TN, true negatives; TP, true positives.")
    ac.font = abbrev_font
    ac.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[abbrev_row].height = 40

    widths = {1: 30,
              2: 18, 3: 14, 4: 14, 5: 19,
              6: 18, 7: 14, 8: 14, 9: 19}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[group_row].height = 26
    ws.row_dimensions[header_row].height = 32

    wb.save(xlsx_path)


# ═════════════════════════════════════════════════════════════════════════
# ─── MASTER ANALYSE-FUNCTIE ──────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════
def run_analysis(results, cohort_name, output_dir, is_normalized=False):
    """Voert de volledige (sex-gestratificeerde) classificatie-analyse uit voor
    één cohort. Identieke statistische logica per cohort; alleen de input-
    DataFrame en output-map verschillen.

    Parameters
    ----------
    results : pd.DataFrame
        Voorspellingen + ECG criteria voor deze cohort (mogelijk gefilterd
        op sex).
    cohort_name : str
        Naam (Totaal / Mannen / Vrouwen / Genormaliseerd).
    output_dir : str
        Sub-map waar figuur, CSV en Excel naartoe geschreven worden.
    is_normalized : bool
        True voor het Genormaliseerd cohort; voegt een extra voetnoot toe.
    """
    print(f"\n\n{'#'*72}")
    print(f"#  ANALYSE VOOR COHORT: {cohort_name}  (n={len(results)})")
    print(f"#  Output directory: {output_dir}")
    print(f"{'#'*72}")

    os.makedirs(output_dir, exist_ok=True)

    if len(results) < 5:
        print(f"  [WAARSCHUWING] Cohort '{cohort_name}' heeft minder dan 5 rijen "
              f"(n={len(results)}). Analyse overgeslagen.")
        return

    results = results.reset_index(drop=True).copy()

    if results["lvh_label"].nunique() < 2:
        print(f"  [WAARSCHUWING] Cohort '{cohort_name}' heeft slechts één LVH-klasse. "
              "ROC/Excel analyse overgeslagen.")
        return

    # Label-mapping voor normalisatie-cohort
    if is_normalized:
        NORM_NOTE = ("Pooled cohort analysis. The LVM-AI receives sex as a model input "
                     "feature, so its P(LVH) output is intrinsically sex-aware. Classical "
                     "ECG criteria retain their sex-specific clinical thresholds (Cornell: "
                     "2.8 mV men / 2.0 mV women; Peguero-Lo Presti: 2.8 mV men / 2.3 mV women).")
    else:
        NORM_NOTE = ""

    # ── Globale drempels (op deze cohort) ───────────────────────────────
    y_true_all  = results["lvh_label"].values.astype(int)
    y_score_all = results["p_lvh"].values.astype(float)
    n_pos_all   = int((y_true_all == 1).sum())
    n_neg_all   = int((y_true_all == 0).sum())

    fpr_g, tpr_g, thr_g = roc_curve(y_true_all, y_score_all)
    youden = tpr_g - fpr_g
    opt_idx = int(np.argmax(youden))
    youden_thr = float(thr_g[opt_idx])

    mask_spec = (1 - fpr_g) >= TARGET_SPEC
    if mask_spec.any():
        idx_spec = int(np.argmax(tpr_g * mask_spec))
        spec95_thr = float(thr_g[idx_spec])
    else:
        spec95_thr = np.nan

    print(f"\n{'='*64}")
    print(f"  GLOBALE LVM-AI DREMPELS (op cohort '{cohort_name}')")
    print(f"{'='*64}")
    print(f"  Optimale Youden drempel : {youden_thr:.4f}")
    spec_thr_str = f"{spec95_thr:.4f}" if not np.isnan(spec95_thr) else "n/a"
    print(f"  95% Spec drempel        : {spec_thr_str}")
    print(f"  Populatie totaal        : {len(results)} (LVH+ n={n_pos_all}, LVH- n={n_neg_all})")

    # ── Stratificatie Male / Female binnen deze cohort ──────────────────
    df_male   = results[results["sex"] == MALE_SEX_VALUE].copy().reset_index(drop=True)
    df_female = results[results["sex"] == FEMALE_SEX_VALUE].copy().reset_index(drop=True)

    print(f"\n  Mannen   : n={len(df_male)}   (LVH+ n={int((df_male['lvh_label']==1).sum())}, "
          f"LVH- n={int((df_male['lvh_label']==0).sum())})")
    print(f"  Vrouwen  : n={len(df_female)}  (LVH+ n={int((df_female['lvh_label']==1).sum())}, "
          f"LVH- n={int((df_female['lvh_label']==0).sum())})")

    male_r   = evaluate_stratum(df_male,   "Male",   youden_thr, spec95_thr)
    female_r = evaluate_stratum(df_female, "Female", youden_thr, spec95_thr)

    print_stratum(f"MALE ATHLETES [{cohort_name}]",   male_r,   youden_thr, spec95_thr)
    print_stratum(f"FEMALE ATHLETES [{cohort_name}]", female_r, youden_thr, spec95_thr)

    # ── ROC figuur ──────────────────────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(15, 7))
    plot_panel(axes[0], male_r,
               f"ROC – Male Athletes (n={male_r['n']}, LVH+ n={male_r['n_pos']})")
    plot_panel(axes[1], female_r,
               f"ROC – Female Athletes (n={female_r['n']}, LVH+ n={female_r['n_pos']})")
    fig.suptitle(f"[{cohort_name}] LVM-AI vs Classical ECG Criteria – Sex-Stratified Performance",
                 fontsize=14)
    if NORM_NOTE:
        fig.text(0.5, 0.01, NORM_NOTE, ha="center", va="bottom",
                 fontsize=8, style="italic", wrap=True)
        fig.tight_layout(rect=[0, 0.05, 1, 0.97])
    else:
        fig.tight_layout()
    fig.savefig(os.path.join(output_dir, "roc_stratified.png"), dpi=150)
    print(f"\nROC figuur opgeslagen: {output_dir}/roc_stratified.png")

    # ── Per-ECG CSV ─────────────────────────────────────────────────────
    cohort_results = results.copy()
    yt_all = cohort_results["lvh_label"].values.astype(int)
    p_all  = cohort_results["p_lvh"].values.astype(float)

    cohort_results["threshold_youden_global"] = youden_thr
    cohort_results["pred_youden"]   = (p_all >= youden_thr).astype(int)
    cohort_results["outcome_youden"]= label_outcome(yt_all, cohort_results["pred_youden"].values)

    if not np.isnan(spec95_thr):
        cohort_results["threshold_spec95_global"] = spec95_thr
        cohort_results["pred_spec95"]   = (p_all >= spec95_thr).astype(int)
        cohort_results["outcome_spec95"]= label_outcome(yt_all, cohort_results["pred_spec95"].values)

    cohort_results["outcome_sokolow"] = label_outcome(yt_all, cohort_results["sokolow_lyon_pred"].values.astype(int))
    cohort_results["outcome_cornell"] = label_outcome(yt_all, cohort_results["cornell_pred"].values.astype(int))
    cohort_results["outcome_peguero"] = label_outcome(yt_all, cohort_results["peguero_pred"].values.astype(int))

    csv_path = os.path.join(output_dir, "classificatie_voorspellingen.csv")
    cohort_results.to_csv(csv_path, index=False)
    print(f"Voorspellingen opgeslagen: {csv_path}")

    # ── Side-by-side Excel ──────────────────────────────────────────────
    xlsx_path = os.path.join(output_dir, "classification_performance.xlsx")
    write_excel(xlsx_path, cohort_name, male_r, female_r,
                youden_thr, spec95_thr, NORM_NOTE=NORM_NOTE)
    print(f"Excel opgeslagen: {xlsx_path}")

    plt.close('all')


# ═════════════════════════════════════════════════════════════════════════
# ─── VIER COHORTEN UITVOEREN ─────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════
cohorts = [
    ("Totaal",  full_results.copy(),                                                                    os.path.join(BASE_OUTPUT_DIR, "Totaal")),
    ("Mannen",  full_results[full_results["sex"] == MALE_SEX_VALUE].reset_index(drop=True).copy(),       os.path.join(BASE_OUTPUT_DIR, "Mannen")),
    ("Vrouwen", full_results[full_results["sex"] == FEMALE_SEX_VALUE].reset_index(drop=True).copy(),     os.path.join(BASE_OUTPUT_DIR, "Vrouwen")),
]

for name, df_cohort, path in cohorts:
    run_analysis(df_cohort, name, path)

# 4e cohort: Genormaliseerd (gepoolde data, sex als model-input vormt de
# intrinsieke sex-normalisatie; klinische ECG-drempels blijven sex-specifiek).
norm_cohort = ("Genormaliseerd", full_results.copy(),
               os.path.join(BASE_OUTPUT_DIR, "Genormaliseerd"))
run_analysis(norm_cohort[1], norm_cohort[0], norm_cohort[2], is_normalized=True)
cohorts.append(norm_cohort)

print("\n" + "="*72)
print("KLAAR. Alle vier cohort-analyses zijn voltooid.")
for name, df_cohort, path in cohorts:
    print(f"  - {name:<16s} (n={len(df_cohort):3d})  ->  {path}/")
print("="*72)
